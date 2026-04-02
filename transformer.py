import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from collections import defaultdict
from datetime import datetime

# ============================================
# CONFIGURATION
# ============================================
INPUT_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "Abdullah", "vertical_data.xlsx")
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUTPUT_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "Abdullah", f"horizontal_output_{timestamp}.xlsx")
PREFERRED_VENDOR = "AMETEK"

# Colors
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def get_priority(company_name, supplier_part, user_part):
    """Determine priority: AMETEK=0, part_match=1, standard=2"""
    company_upper = str(company_name).upper()

    if PREFERRED_VENDOR.upper() in company_upper:
        return 0, "ametek"

    # Check part number match
    if user_part and str(user_part).upper() in str(supplier_part).upper():
        return 1, "part_match"

    return 2, "standard"


def transform_vertical_to_horizontal(input_file, output_file):
    """Transform vertical supplier data to horizontal format"""

    print(f"Reading: {input_file}")
    df = pd.read_excel(input_file)

    # Print columns for debugging
    print(f"Columns found: {list(df.columns)}")
    print(f"Total rows: {len(df)}")

    # Normalize column names
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

    # Try to identify columns
    # Expected: user_nsn, user_part, supplier_part, cage, company
    nsn_col = None
    part_col = None
    supplier_part_col = None
    company_col = None

    for col in df.columns:
        if 'nsn' in col or 'stock' in col:
            nsn_col = col
        elif 'user_part' in col or col == 'user_part':
            part_col = col
        elif 'supplier' in col and 'part' in col:
            supplier_part_col = col
        elif 'company' in col:
            company_col = col

    # If columns not found, use positional
    if not nsn_col:
        nsn_col = df.columns[0]
    if not part_col:
        part_col = df.columns[1] if len(df.columns) > 1 else nsn_col
    if not supplier_part_col:
        supplier_part_col = df.columns[2] if len(df.columns) > 2 else part_col
    if not company_col:
        company_col = df.columns[4] if len(df.columns) > 4 else df.columns[-1]

    print(f"Using columns: NSN={nsn_col}, Part={part_col}, Supplier_Part={supplier_part_col}, Company={company_col}")

    # Group by NSN and User_Part
    nsn_groups = defaultdict(list)

    for _, row in df.iterrows():
        user_nsn = str(row[nsn_col])
        user_part = str(row[part_col]) if part_col in row else ""
        supplier_part = str(row[supplier_part_col]) if supplier_part_col in row else ""
        company = str(row[company_col]) if company_col in row else ""

        priority_num, priority_type = get_priority(company, supplier_part, user_part)

        nsn_groups[(user_nsn, user_part)].append({
            'company': company,
            'part': supplier_part,
            'priority': priority_num,
            'type': priority_type
        })

    # Sort suppliers within each group
    for key in nsn_groups:
        nsn_groups[key].sort(key=lambda x: x['priority'])

    # Find max suppliers
    max_suppliers = max(len(s) for s in nsn_groups.values()) if nsn_groups else 1
    print(f"Max suppliers per NSN: {max_suppliers}")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Horizontal Data"

    # Headers
    headers = ["User_NSN", "User_Part"]
    for i in range(1, max_suppliers + 1):
        headers.extend([f"OEM {i}", f"PN {i}"])

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data
    row_idx = 2
    for (user_nsn, user_part), suppliers in nsn_groups.items():
        # NSN column
        ws.cell(row=row_idx, column=1, value=user_nsn).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).font = DATA_FONT

        # User_Part column
        ws.cell(row=row_idx, column=2, value=user_part).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).font = DATA_FONT

        # Suppliers horizontally
        col_idx = 3
        for supplier in suppliers:
            # OEM (Company)
            cell_oem = ws.cell(row=row_idx, column=col_idx, value=supplier['company'])
            cell_oem.border = THIN_BORDER
            cell_oem.font = DATA_FONT

            if supplier['type'] == "ametek":
                cell_oem.fill = GREEN_FILL
            elif supplier['type'] == "part_match":
                cell_oem.fill = YELLOW_FILL

            # PN (Supplier Part)
            cell_pn = ws.cell(row=row_idx, column=col_idx + 1, value=supplier['part'])
            cell_pn.border = THIN_BORDER
            cell_pn.font = DATA_FONT

            col_idx += 2

        row_idx += 1

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    for i in range(3, len(headers) + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        if (i - 3) % 2 == 0:
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 15

    # Freeze header
    ws.freeze_panes = 'A2'

    # Save
    wb.save(output_file)
    print(f"\n✓ Saved to: {output_file}")
    print(f"✓ Total NSNs: {len(nsn_groups)}")

    return output_file


if __name__ == "__main__":
    print("="*60)
    print("VERTICAL TO HORIZONTAL TRANSFORMER")
    print("="*60)

    if not os.path.exists(INPUT_FILE):
        print(f"\nERROR: Input file not found: {INPUT_FILE}")
        print("\nPlease save your vertical data as:")
        print(f"  {INPUT_FILE}")
        print("\nOr edit INPUT_FILE in this script to point to your file.")
    else:
        transform_vertical_to_horizontal(INPUT_FILE, OUTPUT_FILE)
        os.startfile(OUTPUT_FILE)
