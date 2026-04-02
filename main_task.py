"""
SAMI TURBO SCRAPER - 8 PARALLEL BOTS
BATCH 2 - ITEMS 501-1000, HEADLESS
"""

import subprocess
import sys

for package in ["selenium", "pandas", "openpyxl", "psutil"]:
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package, "-q"])

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import re
import time
import random
import threading
import psutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Master DB - add server/ to path so we can import master_db
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "server"))
import master_db

# ============================================
# CONFIGURATION
# ============================================
OUTPUT_FILE = None        # Set dynamically based on input file
NUM_WORKERS = 8           # Default, overridden by Smart Scaling
SAVE_INTERVAL = 50        # Save every 50 items
PAGE_TIMEOUT = 15         # 15-second timeout for page load
PRICE_LOGIC = "none"      # "high", "low", or "none"

# Smart Scaling
MAX_BOTS = 10             # Hard cap on bots
RAM_PER_BOT_GB = 1.0      # 1 GB free RAM per bot
SMALL_JOB_THRESHOLD = 20  # Items below this = only 2 bots
SMALL_JOB_BOTS = 2
STAGGER_DELAY = (2, 5)    # Seconds between bot launches (min, max) - reduced for speed
COOLDOWN_SECONDS = 15     # Sleep on page-load failure before retry - reduced

# Pre-compiled regex for price/date extraction (faster than re.search with string)
PRICE_RE = re.compile(r'\$[\d,]+\.\d{2}')
DATE_RE = re.compile(r'[A-Z][a-z]{2}-\d{2}-\d{4}')
CAGE_RE = re.compile(r'^[A-Z0-9]{5}$')

BLACK_LIST = frozenset(["A486G", "FINLAND"])
_BAD_PREFIXES = ("HUES", "ABGL", "SHPE", "FSC", "NIIN", "NSN", "MOE", "AAC", "RNCC", "RNVC", "DAC", "RNAAC", "CAGE")
_SKIP_COMPANIES = frozenset(["NATURAL", "BLACK", "RECTANGULAR", "MINIMUM"])

# Styles
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

LOGIQUEST_URL = "https://www.lqlite.com/"

# Thread-safe data storage
data_lock = threading.Lock()
all_results = []
progress_lock = threading.Lock()
worker_progress = {}


def calculate_optimal_bots(total_items):
    """Calculate optimal bot count based on system resources and job size."""
    mem = psutil.virtual_memory()
    free_gb = round(mem.available / (1024 ** 3), 1)
    total_gb = round(mem.total / (1024 ** 3), 1)
    cpu_percent = psutil.cpu_percent(interval=0.5)

    # 1 bot per GB of free RAM, capped at MAX_BOTS
    ram_bots = max(1, int(free_gb / RAM_PER_BOT_GB))
    ram_bots = min(ram_bots, MAX_BOTS)

    # Small jobs: cap at SMALL_JOB_BOTS to save resources
    if total_items < SMALL_JOB_THRESHOLD:
        optimal = min(ram_bots, SMALL_JOB_BOTS)
    else:
        optimal = ram_bots

    # Never more bots than items
    optimal = min(optimal, total_items)
    optimal = max(1, optimal)

    return optimal, free_gb, total_gb, cpu_percent


def create_driver():
    """Create HEADLESS Chrome browser with anti-detection stealth"""
    options = Options()

    # ========================================
    # HEADLESS MODE - ENABLED FOR SPEED
    # ========================================
    options.add_argument("--headless=new")  # ENABLED FOR PRODUCTION

    # Basic stability options
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--log-level=3")

    # Anti-bot detection
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36")
    options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)

    # Desktop window size
    options.add_argument("--window-size=1920,1080")

    # Disable images & CSS for faster page loads
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--blink-settings=imagesEnabled=false")

    # PAGE LOAD STRATEGY: 'eager' = don't wait for background scripts
    options.page_load_strategy = 'eager'

    driver = webdriver.Chrome(options=options)

    # Remove navigator.webdriver flag via CDP
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })

    # Set referer header to mimic traffic from Google
    driver.execute_cdp_cmd("Network.enable", {})
    driver.execute_cdp_cmd("Network.setExtraHTTPHeaders", {
        "headers": {"Referer": "https://www.google.com/"}
    })

    driver.set_page_load_timeout(30)
    driver.set_script_timeout(PAGE_TIMEOUT)
    # NOTE: implicitly_wait removed - conflicts with explicit WebDriverWait and adds hidden delays
    return driver


def is_valid_stock_number(value):
    if not value:
        return False
    clean_val = str(value).strip().replace("-", "").replace(" ", "")
    return clean_val.isdigit() and 12 <= len(clean_val) <= 14


def load_stock_numbers(input_file):
    """Load all stock numbers from the given Excel file"""
    if not os.path.exists(input_file):
        print(f"ERROR: File not found: {input_file}")
        return []

    print(f"Reading: {input_file}")

    try:
        xl_file = pd.ExcelFile(input_file)

        for sheet_name in xl_file.sheet_names:
            df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str, header=None)
            if df.empty:
                continue

            for col_idx in range(min(5, len(df.columns))):
                for row_idx in range(len(df)):
                    cell_value = df.iloc[row_idx, col_idx]
                    if is_valid_stock_number(cell_value):
                        stock_numbers = []
                        for r in range(row_idx, len(df)):
                            val = df.iloc[r, col_idx]
                            if is_valid_stock_number(val):
                                clean = str(val).strip().replace("-", "").replace(" ", "")
                                stock_numbers.append(clean)

                        if stock_numbers:
                            return stock_numbers
        return []
    except Exception as e:
        print(f"Error: {e}")
        return []


def search_stock(driver, wait, stock_number, worker_id):
    """Search single stock - VISIBLE mode with LOGGING"""

    # ========================================
    # CLEAN INPUT - Remove hidden spaces
    # ========================================
    clean_stock = stock_number.strip()

    # Reduced logging for production (no per-item logs)

    try:
        # Try to find search box; if missing, reload main page with smart wait
        try:
            search_box = wait.until(EC.presence_of_element_located((By.ID, "nALL")))
        except Exception as e:
            print(f"   [BOT {worker_id}] Search box not found, reloading main page...")
            driver.get(LOGIQUEST_URL)
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            search_box = wait.until(EC.presence_of_element_located((By.ID, "nALL")))

        # Clear and enter stock number (use JS for speed)
        driver.execute_script("arguments[0].value = '';", search_box)
        search_box.send_keys(clean_stock + Keys.RETURN)

        # Smart wait: wait for page content to change instead of fixed sleep
        time.sleep(1.5)

        # ========================================
        # MULTI-RESULT HANDLING: Click first result if on list page
        # ========================================
        # Use JS to check page text (faster than fetching full page_source)
        has_results = driver.execute_script(
            "var t = document.body.innerText; return t.indexOf('Search Results:') !== -1 || t.toLowerCase().indexOf('results found') !== -1;"
        )
        if has_results:
            # We're on a multi-result page - need to click the first NIIN link
            try:
                # Look for clickable NIIN links (usually in a table with class or in specific format)
                niin_links = driver.find_elements(By.XPATH, "//a[contains(@href, 'NIIN') or contains(@href, 'niin')]")
                if not niin_links:
                    # Try finding links that look like NIIN (9-digit numbers)
                    niin_links = driver.find_elements(By.XPATH, "//a[string-length(normalize-space(text()))=9 and translate(text(),'0123456789','')='']")
                if not niin_links:
                    # Try finding any link in the results table
                    niin_links = driver.find_elements(By.XPATH, "//table//tr//td//a")

                if niin_links:
                    # Click the first valid NIIN link
                    for link in niin_links:
                        link_text = link.text.strip()
                        if not link_text or len(link_text) < 5:
                            continue
                        link.click()
                        time.sleep(1.5)
                        break
            except Exception as click_err:
                pass  # Continue with current page if click fails

        all_rows = driver.find_elements(By.TAG_NAME, "tr")
        final_stock = ""
        niin = ""
        raw_data = []

        # Extract FSC and NIIN for stock number
        for row in all_rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 2:
                texts = [c.text.strip() for c in cells]
                if texts[0] == "NIIN:" and len(texts) > 1:
                    niin = texts[1]
                if texts[0] == "FSC:" and len(texts) > 1:
                    fsc = texts[1]
                    if niin:
                        final_stock = f"{fsc}{niin}"

        # ========================================
        # EXTRACT PRICE DATA (based on PRICE_LOGIC)
        # Only from the MANAGEMENT table (has "Action Date" header)
        # ========================================
        # Scroll to bottom to ensure Management table is loaded
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.5)
        all_rows = driver.find_elements(By.TAG_NAME, "tr")

        mgmt_data = {"Action Date": "", "UI": "", "Unit Price": ""}
        if PRICE_LOGIC != "none":
            try:
                in_mgmt_table = False
                price_entries = []

                for mr in all_rows:
                    row_text = mr.text.strip().upper()

                    # Detect management table header (may span multiple rows)
                    if not in_mgmt_table:
                        if "ACTION DATE" in row_text or "MANAGEMENT INFORMATION" in row_text:
                            in_mgmt_table = True
                        continue

                    mcells = mr.find_elements(By.TAG_NAME, "td")
                    if len(mcells) < 3:
                        # Skip sub-header rows but don't break yet
                        continue

                    mt = [c.text.strip() for c in mcells]
                    row_joined = " ".join(mt)

                    # Must have a price - if no price and we already have entries, table ended
                    if not PRICE_RE.search(row_joined):
                        if price_entries:
                            break
                        continue
                    if not DATE_RE.search(row_joined):
                        continue

                    price_entries.append(mt)

                if price_entries:
                    parsed = []
                    for pe in price_entries:
                        row_price = None
                        row_date = None
                        row_ui = ""
                        for cell in pe:
                            pm = PRICE_RE.search(cell)
                            if pm and row_price is None:
                                try:
                                    row_price = float(pm.group(0).replace("$", "").replace(",", ""))
                                except ValueError:
                                    pass
                            dm = DATE_RE.search(cell)
                            if dm and row_date is None:
                                try:
                                    row_date = datetime.strptime(dm.group(0), "%b-%d-%Y")
                                except ValueError:
                                    pass
                            if len(cell) == 2 and cell.isalpha() and cell.isupper() and not row_ui:
                                row_ui = cell
                        if row_price is not None and row_date is not None:
                            parsed.append({"cells": pe, "price": row_price, "date": row_date, "ui": row_ui})

                    if parsed:
                        if PRICE_LOGIC == "high":
                            best = max(parsed, key=lambda x: (x["date"], x["price"]))
                        else:  # "low"
                            best = max(parsed, key=lambda x: (x["date"], -x["price"]))

                        mgmt_data["Unit Price"] = f"${best['price']:,.2f}"
                        mgmt_data["Action Date"] = best["date"].strftime("%b-%d-%Y")
                        if best["ui"]:
                            mgmt_data["UI"] = best["ui"]
            except Exception:
                pass

        # Find all rows in all tables (reusing all_rows from above)
        for row in all_rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 3:
                texts = [c.text.strip() for c in cells]

                # Skip rows with metadata labels
                first_cell = texts[0].upper() if texts else ""
                if any(skip in first_cell for skip in ["NIIN", "FSC", "NSN", "MOE", "AAC", ":"]):
                    continue

                # Skip blacklisted
                if any(word in first_cell for word in BLACK_LIST):
                    continue

                # Find CAGE code position (5 alphanumeric)
                cage_idx = -1
                for idx, txt in enumerate(texts):
                    if txt and len(txt) == 5 and CAGE_RE.match(txt):
                        cage_idx = idx
                        break

                if cage_idx > 0:  # CAGE found and not in first position
                    part_num = texts[0].strip()

                    # SKIP bad part numbers
                    if any(part_num.upper().startswith(bp) for bp in _BAD_PREFIXES):
                        continue
                    if len(part_num) <= 3:
                        continue

                    # Get company name - look in remaining cells
                    company_name = ""
                    for idx in range(cage_idx + 1, min(cage_idx + 4, len(texts))):
                        candidate = texts[idx].strip()
                        if len(candidate) <= 5:
                            continue
                        if candidate.upper() in _SKIP_COMPANIES:
                            continue
                        if "INCH" in candidate.upper():
                            continue
                        company_name = candidate
                        break

                    if part_num and company_name and len(company_name) > 5:
                        raw_data.append([part_num, texts[cage_idx], company_name])

        if not final_stock:
            final_stock = clean_stock
        final_stock = str(final_stock).replace("-", "")

        # ========================================
        # SEPARATE AMETEK FROM OTHER COMPANIES
        # ========================================
        ametek_entries = []
        other_entries = []
        used_entries = set()  # Track (part_num, company) pairs, not just part_num

        for row in raw_data:
            part_num = row[0].strip() if row else ""
            company = row[2].strip() if len(row) > 2 else ""
            company_upper = company.upper()

            entry_key = (part_num, company_upper)
            if not part_num or entry_key in used_entries:
                continue

            # Skip bad entries
            skip_words = ["HUES", "ABGL", "SHPE", "CRF,", "NATURAL", "BLACK", "RECTANGULAR", "FSC", "NIIN"]
            if any(sw in company_upper for sw in skip_words):
                continue
            if len(company) <= 5:
                continue

            used_entries.add(entry_key)

            # STRICT AMETEK CHECK
            if "AMETEK" in company_upper:
                ametek_entries.append((part_num, "AMETEK"))
            else:
                other_entries.append((part_num, company))

        result = {
            "Stock Number": final_stock,
            "Action Date": mgmt_data["Action Date"],
            "UI": mgmt_data["UI"],
            "Unit Price": mgmt_data["Unit Price"],
        }

        # ========================================
        # MANDATORY AMETEK SLOT (Col B & C)
        # ========================================
        if ametek_entries:
            # AMETEK FOUND - fill slot 1
            result["P.NO 1"] = ametek_entries[0][0]
            result["MFG 1"] = "AMETEK"
            ametek_entries = ametek_entries[1:]
        else:
            # AMETEK NOT FOUND - LEAVE SLOT 1 BLANK
            result["P.NO 1"] = ""
            result["MFG 1"] = ""

        # ========================================
        # HORIZONTAL LAYOUT - Other MFGs from Col D onwards
        # ========================================
        col_idx = 2  # Start from P.NO 2 (Column D)

        # Any additional AMETEK entries
        for part_num, mfg in ametek_entries:
            result[f"P.NO {col_idx}"] = part_num
            result[f"MFG {col_idx}"] = mfg
            col_idx += 1

        # All other manufacturers
        for part_num, mfg in other_entries:
            result[f"P.NO {col_idx}"] = part_num
            result[f"MFG {col_idx}"] = mfg
            col_idx += 1

        # Success logged via progress counter
        return result, "success"

    except Exception as e:
        # ========================================
        # ERROR HANDLING - Log but don't crash
        # ========================================
        error_msg = str(e).lower()
        if "timeout" in error_msg or "timed out" in error_msg:
            print(f"   [BOT {worker_id}] TIMEOUT: {clean_stock} - cooling down {COOLDOWN_SECONDS}s...")
            time.sleep(COOLDOWN_SECONDS)
        elif "session" in error_msg or "invalid session" in error_msg:
            print(f"   [BOT {worker_id}] SESSION DEAD: {clean_stock}")
            return None, "session_dead"
        elif "no such element" in error_msg:
            print(f"   [BOT {worker_id}] ELEMENT NOT FOUND: {clean_stock}")
        else:
            print(f"   [BOT {worker_id}] ERROR: {clean_stock} -> {str(e)[:50]}")
            time.sleep(COOLDOWN_SECONDS)

        # Return empty result instead of crashing
        return {"Stock Number": clean_stock, "Action Date": "", "UI": "", "Unit Price": "",
                "P.NO 1": "", "MFG 1": ""}, "error"


def worker_scrape(worker_id, stock_chunk, stagger=True):
    """Worker function - runs in separate thread"""
    global all_results, worker_progress

    # Staggered Start: delay launch to avoid IP rate limits
    if stagger and worker_id > 1:
        delay = (worker_id - 1) * random.uniform(*STAGGER_DELAY)
        print(f"   [BOT {worker_id}] Staggered start: waiting {delay:.0f}s...")
        time.sleep(delay)

    driver = None
    wait = None
    local_results = []
    success = 0
    errors = 0
    no_data = 0
    restarts = 0

    def start_browser():
        nonlocal driver, wait
        try:
            if driver:
                driver.quit()
        except:
            pass
        driver = create_driver()
        wait = WebDriverWait(driver, PAGE_TIMEOUT)

        # Load main page with smart wait and retry
        for attempt in range(3):
            try:
                driver.get(LOGIQUEST_URL)
                # Smart wait: ensure page body is fully loaded
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                # Wait for the search box to appear (confirms page is interactive)
                search_box = wait.until(EC.presence_of_element_located((By.ID, "nALL")))
                if search_box:
                    print(f"\n   [BOT {worker_id}] Browser ready!")
                    return True
            except Exception as e:
                if attempt < 2:
                    print(f"   [BOT {worker_id}] Page load attempt {attempt+1} failed, cooling down {COOLDOWN_SECONDS}s...")
                    time.sleep(COOLDOWN_SECONDS)
                else:
                    print(f"   [BOT {worker_id}] ERROR: Could not load page after 3 attempts")
        return False

    if not start_browser():
        return worker_id, 0, len(stock_chunk), 0

    try:
        for i, stock in enumerate(stock_chunk):
            # Update progress
            with progress_lock:
                worker_progress[worker_id] = {
                    "done": success,
                    "total": len(stock_chunk),
                    "errors": errors,
                    "no_data": no_data
                }

            result, status = search_stock(driver, wait, stock, worker_id)

            if status == "success" and result:
                if result.get("Stock Number", "").strip():
                    local_results.append(result)
                    success += 1

                    # Add to global results periodically (+ batch DB save)
                    if success % 10 == 0:
                        with data_lock:
                            all_results.extend(local_results)
                            # Batch save to DB (faster than per-item)
                            try:
                                master_db.upsert_batch(local_results, logic_used=PRICE_LOGIC)
                            except Exception:
                                pass
                            local_results = []

            elif status == "session_dead":
                restarts += 1
                if restarts > 10:
                    print(f"   [BOT {worker_id}] Too many restarts, stopping.")
                    break
                print(f"   [BOT {worker_id}] Session dead, cooling down {COOLDOWN_SECONDS}s before restart (#{restarts})...")
                time.sleep(COOLDOWN_SECONDS)
                start_browser()

            elif status == "error":
                errors += 1
                # Still save the partial result
                if result:
                    local_results.append(result)

            else:
                no_data += 1

    except Exception as e:
        print(f"\n   [BOT {worker_id}] CRITICAL ERROR: {e}")

    finally:
        try:
            driver.quit()
        except:
            pass

    # Add remaining results + batch DB save
    with data_lock:
        all_results.extend(local_results)
        if local_results:
            try:
                master_db.upsert_batch(local_results, logic_used=PRICE_LOGIC)
            except Exception:
                pass

    return worker_id, success, errors, no_data


def save_to_excel(data_rows):
    """Save sorted by Stock Number with AMETEK green highlighting"""
    if not data_rows:
        return 0, 0

    # Sort by Stock Number ascending
    sorted_rows = sorted(data_rows, key=lambda r: str(r.get("Stock Number", "")))

    # Count AMETEK rows
    ametek_rows = [r for r in sorted_rows if str(r.get("MFG 1", "")).strip().upper() == "AMETEK"]

    wb = Workbook()
    ws = wb.active
    ws.title = "SAMI Verify Logic"

    # Build headers
    all_columns = set()
    for row in sorted_rows:
        all_columns.update(row.keys())

    headers = ["Stock Number", "Action Date", "UI", "Unit Price"]
    max_supplier = 1
    for col in all_columns:
        if col.startswith("P.NO ") or col.startswith("MFG "):
            try:
                num = int(col.split()[-1])
                if num > max_supplier:
                    max_supplier = num
            except:
                pass

    for i in range(1, max_supplier + 1):
        headers.append(f"P.NO {i}")
        headers.append(f"MFG {i}")

    # Drop entirely empty columns
    non_empty = []
    for h in headers:
        if h in ("Stock Number", "Action Date", "UI", "Unit Price"):
            non_empty.append(h)
        else:
            if any(r.get(h, "") for r in sorted_rows):
                non_empty.append(h)
    headers = non_empty

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    ametek_count = 0
    for row_idx, row_data in enumerate(sorted_rows, 2):
        has_ametek = any(
            "AMETEK" in str(row_data.get(k, "")).upper()
            for k in row_data if k.startswith("MFG ")
        )
        if has_ametek:
            ametek_count += 1

        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.fill = WHITE_FILL

            # Green only on AMETEK MFG cell + its paired P.NO cell
            if header.startswith("MFG ") and value and "AMETEK" in str(value).upper():
                cell.fill = GREEN_FILL
                slot = header.split()[-1]
                pno_header = f"P.NO {slot}"
                if pno_header in headers:
                    pno_ci = headers.index(pno_header) + 1
                    ws.cell(row=row_idx, column=pno_ci).fill = GREEN_FILL

            if col_idx == 1:
                cell.number_format = '@'
                cell.value = str(value).strip() if value else ""

    # Auto-fit
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT_FILE)
    wb.close()

    return ametek_count, len(ametek_rows)


def print_progress():
    """Print progress for all workers"""
    with progress_lock:
        total_done = sum(w.get("done", 0) for w in worker_progress.values())
        total_target = sum(w.get("total", 0) for w in worker_progress.values())

        status = f"\r[TOTAL: {total_done:,}/{total_target:,}] "
        for wid in sorted(worker_progress.keys()):
            w = worker_progress[wid]
            status += f"BOT{wid}:{w.get('done',0)}/{w.get('total',0)} "

        print(status, end="", flush=True)
        return total_done


def main():
    global all_results, OUTPUT_FILE, PRICE_LOGIC, NUM_WORKERS

    print("\n" + "=" * 70)
    print("   SAMI TURBO SCRAPER - SMART SCALING")
    print("=" * 70)

    # Initialize master database
    master_db.init_master_db()

    # Clear all_results for fresh start
    with data_lock:
        all_results = []

    # Step 1: Ask user for input file
    while True:
        raw_path = input("\n   Please drag and drop your Excel file here and press Enter: ").strip()
        # Clean quotes that Windows adds on drag & drop
        input_file = raw_path.strip('"').strip("'").strip()
        if not input_file:
            print("   ERROR: No file provided. Please try again.")
            continue
        if not os.path.isfile(input_file):
            print("   File not found! Please try again.")
            continue
        if not input_file.lower().endswith(('.xlsx', '.xls')):
            print("   File not found! Please try again.")
            continue
        break

    # Set output file based on input file name
    input_dir = os.path.dirname(input_file)
    input_name = os.path.splitext(os.path.basename(input_file))[0]
    OUTPUT_FILE = os.path.join(input_dir, f"3LINES_{input_name}_Result.xlsx")

    # Step 2: Load and count all stock numbers
    stock_numbers = load_stock_numbers(input_file)
    if not stock_numbers:
        print("\nERROR: No stock numbers found!")
        return

    total_in_file = len(stock_numbers)
    print(f"\n   Total stock numbers found in Excel: {total_in_file:,}")
    print(f"   Output: {os.path.basename(OUTPUT_FILE)}")
    print("=" * 70)

    # Step 2: Ask user how many to scrape
    while True:
        user_input = input(f"\n   How many items would you like to scrape from this total? (Press Enter for all): ").strip()
        if user_input == "":
            scrape_count = total_in_file
            print(f"   -> Scraping ALL {scrape_count:,} items")
            break
        try:
            scrape_count = int(user_input)
            if scrape_count <= 0:
                print(f"   ERROR: Please enter a number greater than 0.")
                continue
            if scrape_count > total_in_file:
                print(f"   ERROR: You entered {scrape_count:,} but only {total_in_file:,} items exist. Try again.")
                continue
            print(f"   -> Scraping {scrape_count:,} out of {total_in_file:,} items")
            break
        except ValueError:
            print(f"   ERROR: Invalid input. Please enter a number or press Enter for all.")

    # Step 4: Ask for Price Logic
    while True:
        price_input = input("\n   Select Price Logic (1: High Price [Default], 2: Low Price, n: None): ").strip().lower()
        if price_input == "1" or price_input == "":
            PRICE_LOGIC = "high"
            print("   -> Price Logic: HIGH PRICE (most expensive record)")
            break
        elif price_input == "2":
            PRICE_LOGIC = "low"
            print("   -> Price Logic: LOW PRICE (cheapest record)")
            break
        elif price_input == "n":
            PRICE_LOGIC = "none"
            print("   -> Price Logic: NONE (skip price extraction)")
            break
        else:
            print("   ERROR: Invalid input. Enter 1, 2, or n.")

    # Apply limit
    stock_numbers = stock_numbers[:scrape_count]

    # Smart skip: remove stock numbers already in master DB
    try:
        existing = {p["part_number"] for p in master_db.get_all_parts()}
        before = len(stock_numbers)
        stock_numbers = [s for s in stock_numbers if s not in existing]
        skipped = before - len(stock_numbers)
        if skipped > 0:
            print(f"\n   [SMART SKIP] {skipped:,} items already in database - skipping them")
    except Exception:
        pass

    total_stocks = len(stock_numbers)
    if total_stocks == 0:
        print("\n   All items already scraped! Nothing to do.")
        return
    print(f"   Items to scrape: {total_stocks:,}")

    # ============================================
    # SMART SCALING - Calculate optimal bots
    # ============================================
    optimal, free_gb, total_gb, cpu_pct = calculate_optimal_bots(total_stocks)

    print("\n" + "-" * 70)
    print(f"   [SYSTEM] RAM: {free_gb} GB free / {total_gb} GB total  |  CPU: {cpu_pct}%")
    if total_stocks < SMALL_JOB_THRESHOLD:
        print(f"   [SYSTEM] Small job ({total_stocks} items < {SMALL_JOB_THRESHOLD}), capped at {SMALL_JOB_BOTS} bots")
    print(f"   [SYSTEM] Optimal bots calculated: {optimal}")
    print("-" * 70)

    while True:
        confirm = input(f"\n   Start with {optimal} bots? (Y/n, or enter a number): ").strip().lower()
        if confirm in ("", "y", "yes"):
            NUM_WORKERS = optimal
            break
        elif confirm == "n":
            while True:
                custom = input("   Enter number of bots (1-10): ").strip()
                try:
                    custom_n = int(custom)
                    if 1 <= custom_n <= MAX_BOTS:
                        NUM_WORKERS = min(custom_n, total_stocks)
                        break
                    print(f"   ERROR: Must be between 1 and {MAX_BOTS}.")
                except ValueError:
                    print("   ERROR: Enter a valid number.")
            break
        else:
            try:
                custom_n = int(confirm)
                if 1 <= custom_n <= MAX_BOTS:
                    NUM_WORKERS = min(custom_n, total_stocks)
                    break
                print(f"   ERROR: Must be between 1 and {MAX_BOTS}.")
            except ValueError:
                print("   ERROR: Enter Y, n, or a number.")

    print(f"\n   -> Using {NUM_WORKERS} bot(s)")
    print(f"   -> Staggered start: {STAGGER_DELAY[0]}-{STAGGER_DELAY[1]}s between bots")
    print(f"   -> Cooldown on error: {COOLDOWN_SECONDS}s")

    # Split into chunks for each worker (fair distribution)
    avg = max(1, total_stocks // NUM_WORKERS)
    rem = total_stocks % NUM_WORKERS
    chunks = []
    start = 0
    for i in range(NUM_WORKERS):
        extra = 1 if i < rem else 0
        end = start + avg + extra
        if start < total_stocks:
            chunks.append(stock_numbers[start:end])
            print(f"   BOT {i+1}: {start:,} to {end:,} ({end-start:,} items)")
        start = end

    print("\n" + "-" * 70)
    print(f"STARTING SCRAPE ({total_stocks:,} items) with {NUM_WORKERS} bot(s)...")
    print(f"   URL: {LOGIQUEST_URL[:50]}...")
    print("-" * 70)
    print("\n   [INFO] Running in HEADLESS mode with STAGGERED START...")
    print("   [INFO] Progress updates will appear below...")
    print("")

    start_time = time.time()
    last_save = 0

    # Start workers (staggered launch handled inside worker_scrape)
    with ThreadPoolExecutor(max_workers=NUM_WORKERS) as executor:
        futures = {
            executor.submit(worker_scrape, i+1, chunk, True): i+1
            for i, chunk in enumerate(chunks)
        }

        try:
            # Monitor progress
            while any(not f.done() for f in futures):
                time.sleep(2)
                total_done = print_progress()

                # Auto-save
                if total_done - last_save >= SAVE_INTERVAL:
                    with data_lock:
                        if all_results:
                            save_to_excel(all_results)
                            print(f"\n   [AUTO-SAVED] {len(all_results):,} rows")
                            last_save = total_done

            # Get final results
            for future in as_completed(futures):
                wid, success, errors, no_data = future.result()
                print(f"\n   [BOT {wid} DONE] Success:{success} Errors:{errors} NoData:{no_data}")

        except KeyboardInterrupt:
            print("\n\n" + "!" * 70)
            print("   CTRL+C - EMERGENCY SAVE...")
            print("!" * 70)

    # Final save
    with data_lock:
        if all_results:
            ametek_count, ametek_rows = save_to_excel(all_results)

            # Final batch save to master DB (catches any missed by instant save)
            try:
                db_count = master_db.upsert_batch(all_results, logic_used=PRICE_LOGIC)
                print(f"\n   [DB] Master database updated: {db_count:,} parts saved/updated")
            except Exception as db_err:
                print(f"\n   [DB] Warning: Could not save to master DB: {db_err}")

            elapsed = time.time() - start_time
            hours = int(elapsed // 3600)
            mins = int((elapsed % 3600) // 60)
            secs = int(elapsed % 60)

            # DB stats
            try:
                stats = master_db.get_stats()
                db_total = stats["total_parts"]
            except Exception:
                db_total = "?"

            print("\n\n" + "=" * 70)
            print("   SCRAPE COMPLETE!")
            print("=" * 70)
            print(f"   File: {os.path.basename(OUTPUT_FILE)}")
            print(f"   TOTAL Rows: {len(all_results):,}")
            print(f"   AMETEK Rows (GREEN): {ametek_rows:,}")
            print(f"   Master DB Total: {db_total} parts")
            print(f"   Time: {hours}h {mins}m {secs}s")
            print("=" * 70)

            print("\nOpening Excel file...")
            os.startfile(OUTPUT_FILE)
        else:
            print("\n\nNo data collected!")

    print("\n[AUTO-CLOSE] Script completed.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n\nCRITICAL ERROR: {e}")
        # Emergency save
        with data_lock:
            if all_results:
                save_to_excel(all_results)
                print(f"Emergency saved {len(all_results)} rows")
        print("\n[AUTO-CLOSE] Script ended due to error.")
