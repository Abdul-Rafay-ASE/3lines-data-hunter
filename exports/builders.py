"""Excel / CSV / JSON exporters for scraped result rows.

build_excel / build_csv / build_json produce the three download formats
shown on the Scraper tab's completion screen. build_failed_excel
produces the single-column XLSX of stocks that did not produce a result
(used by the Download Failed Stocks button).

All four functions:
  - filter out rows that match `blacklisted_companies` (via
    utils.parsing.row_is_blacklisted)
  - sort priority-matching rows to the top (via row_has_priority)
  - return bytes ready for a Streamlit download_button (or None if the
    input is empty / fully filtered out)

No Streamlit / DB / Selenium coupling.
"""
import io
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from utils.parsing import row_has_priority, row_is_blacklisted


# ── Excel cell styling shared by build_excel + build_failed_excel ──
H_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
H_FONT = Font(bold=True, size=11, color="FFFFFF")
PRIORITY_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
W_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BDR = Border(left=Side(style='thin'), right=Side(style='thin'),
             top=Side(style='thin'), bottom=Side(style='thin'))


def build_excel(data, priority_targets, blacklisted_companies):
    if not data:
        return None, 0, 0, 0
    clean_data, excluded = [], 0
    for r in data:
        if row_is_blacklisted(r, blacklisted_companies):
            excluded += 1
        else:
            clean_data.append(r)
    if not clean_data:
        return None, 0, 0, excluded
    prio_rows = [r for r in clean_data if row_has_priority(r, priority_targets)]
    other_rows = [r for r in clean_data if not row_has_priority(r, priority_targets)]
    rows = prio_rows + other_rows
    wb = Workbook(); ws = wb.active; ws.title = "3LINES Output"
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    mx = max((int(c.split()[-1]) for c in all_keys if c.startswith("P.NO ") or c.startswith("MFG ")), default=1)
    # Unit Price + Action Date come from the Management Information table.
    # They sit between Stock Number and the P.NO/MFG chain so operators
    # see price next to the stock they queried. Old rows without these
    # keys render as empty cells via dict.get(h, "") in the value loop.
    headers = ["Stock Number", "Unit Price", "Action Date"]
    for i in range(1, mx + 1):
        headers += [f"P.NO {i}", f"MFG {i}"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.border = BDR
        cell.alignment = Alignment(horizontal='center', vertical='center')
    priority_count = 0
    for ri, rd in enumerate(rows, 2):
        is_prio = row_has_priority(rd, priority_targets)
        if is_prio:
            priority_count += 1
        for ci, h in enumerate(headers, 1):
            v = rd.get(h, ""); cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = BDR; cell.font = Font(size=10)
            cell.fill = PRIORITY_FILL if is_prio else W_FILL
            if ci == 1:
                cell.number_format = '@'; cell.value = str(v).strip() if v else ""
    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 45)
    ws.column_dimensions['A'].width = 20; ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); wb.close(); buf.seek(0)
    return buf.getvalue(), priority_count, len(rows), excluded


def build_csv(data, pt, bl):
    if not data:
        return None
    cd = [r for r in data if not row_is_blacklisted(r, bl)]
    if not cd:
        return None
    rows = [r for r in cd if row_has_priority(r, pt)] + [r for r in cd if not row_has_priority(r, pt)]
    df = pd.DataFrame(rows); cols = ["Stock Number"] + [c for c in df.columns if c != "Stock Number"]
    return df[cols].to_csv(index=False).encode("utf-8")


def build_json(data, pt, bl):
    if not data:
        return None
    cd = [r for r in data if not row_is_blacklisted(r, bl)]
    if not cd:
        return None
    rows = [r for r in cd if row_has_priority(r, pt)] + [r for r in cd if not row_has_priority(r, pt)]
    return json.dumps(rows, indent=2, ensure_ascii=False).encode("utf-8")


def build_failed_excel(failed_stocks):
    """Single-column XLSX listing the stocks that did not produce a result
    (after the retry pass). Returns bytes, or None if there's nothing to export."""
    if not failed_stocks:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Failed Stocks"
    cell = ws.cell(row=1, column=1, value="Stock Number")
    cell.font = H_FONT; cell.fill = H_FILL; cell.border = BDR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for i, stk in enumerate(failed_stocks, 2):
        c = ws.cell(row=i, column=1, value=str(stk))
        c.border = BDR; c.font = Font(size=10); c.number_format = '@'
    ws.column_dimensions['A'].width = 22
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); wb.close(); buf.seek(0)
    return buf.getvalue()
