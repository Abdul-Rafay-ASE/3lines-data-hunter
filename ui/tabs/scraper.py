"""Scraper tab — upload, speed presets, start button, live progress, completion screen.

The largest tab by far. Owns:
- The Resume / Discard banner for unfinished runs.
- The 3-step indicator at the top.
- Upload, priority/blacklist, smart-skip checkbox.
- Speed-preset buttons (capped at DH_MAX_BOTS).
- The Start button that invokes run_scraper.
- The live tracker placeholders that run_scraper updates.
- The completion screen with Excel/CSV/JSON + failed-stocks downloads.
- The Run History expander.
"""
import base64
import io
import traceback
from datetime import datetime

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook

from config import DEFAULT_URL, DH_MAX_BOTS, MINUTES_PER_ITEM_MANUAL
from database.db import (
    db_discard_run,
    db_get_failed_stocks_for_run_ids,
    db_get_latest_run_id_for_save_name,
    db_get_results_for_run_ids,
    db_get_run_progress_stocks,
    db_get_unfinished_runs,
)
from exports.builders import build_csv, build_excel, build_failed_excel, build_json
from scraper.orchestrator import run_scraper
from ui.components import render_log, rmetric
from utils.logger import logger
from utils.parsing import load_stocks_strict, parse_comma_list


# ──────────────────────────────────────────────────────────
#  Phase B large-file batching helpers
# ──────────────────────────────────────────────────────────
BATCH_SIZE_DEFAULT = 500
BATCH_THRESHOLD = 1000  # rows above this trigger the batch recommendation


def _build_batch_plan(stocks, batch_size, base_name, group_id):
    """Slice `stocks` into batches of size `batch_size`. Each entry is a dict:
    {index (1-based), total, stocks (list slice), save_name}. The last batch
    picks up the remainder."""
    total = (len(stocks) + batch_size - 1) // batch_size
    plan = []
    for i in range(total):
        start = i * batch_size
        end = min(start + batch_size, len(stocks))
        plan.append({
            "index": i + 1,
            "total": total,
            "stocks": stocks[start:end],
            "save_name": f"{base_name}_batch_{i+1}_of_{total}_{group_id}",
        })
    return plan


def _make_batch_excel_bytes(stocks):
    """Build an in-memory .xlsx (Column A, rows 2+) containing the given
    stocks. Used so run_scraper's existing load_stocks_strict path works
    unchanged — we just feed it a synthesized file per batch."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Stock Number")
    for i, s in enumerate(stocks, 2):
        ws.cell(row=i, column=1, value=s)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Per-run session-state keys that should reset BEFORE each batch starts.
# Excludes batch_* keys (we want to preserve the batch group plan) and
# stop_flag (it's a threading.Event reused across batches).
_PER_RUN_RESET_KEYS = [
    "running", "completed", "stopped",
    "output_bytes", "output_name", "autosave_bytes", "autosave_name",
    "processed", "target", "priority_matches", "blacklisted", "errors",
    "start_time", "elapsed", "final_log",
    "perf_data", "failed_stocks", "final_data", "auto_downloaded",
    "smart_skipped",
]


def _reset_per_run_state(defaults):
    """Reset only the per-run keys, preserving batch group state."""
    ss = st.session_state
    for k in _PER_RUN_RESET_KEYS:
        if k in defaults:
            ss[k] = defaults[k]
    ss.stop_flag.clear()
from utils.system import SMART_LIMIT


def render(defaults, colors, selenium_ok):
    _muted = colors["muted"]
    _accent2 = colors["accent2"]
    ss = st.session_state

    # ── Unfinished-run banner ──
    # Surface any in_progress run older than 5 minutes so a crashed/closed
    # session can be resumed. No auto-resume — user must click Resume to opt in.
    if not ss.running:
        try:
            _unfinished = db_get_unfinished_runs(min_age_minutes=5)
        except Exception:
            logger.exception("Failed to query unfinished runs")
            _unfinished = []
        for _ur in _unfinished:
            _ur_id = _ur["run_id"]
            # runs.processed is only written on finalize, so it stays 0 for an
            # interrupted run. Read the live count from run_progress instead.
            try:
                _processed = len(db_get_run_progress_stocks(_ur_id))
            except Exception:
                logger.exception("Failed to count run_progress rows for %s", _ur_id)
                _processed = _ur.get("processed", 0) or 0
            _total = _ur.get("total_stocks", 0) or 0
            _name = _ur.get("save_name") or "(untitled)"
            _when = (_ur.get("created_at") or "")[:16].replace("T", " ")
            with st.container():
                st.warning(
                    f"**Unfinished run detected:** `{_name}` — "
                    f"{_processed:,}/{_total:,} stocks recorded, started {_when}. "
                    "Re-upload the original file and click Resume to continue, or Discard to dismiss."
                )
                _bc1, _bc2, _ = st.columns([1, 1, 6])
                with _bc1:
                    if st.button("Resume", key=f"resume_{_ur_id}", use_container_width=True):
                        ss.resume_run_id = _ur_id
                        logger.info("User clicked Resume for run %s", _ur_id)
                        st.rerun()
                with _bc2:
                    if st.button("Discard", key=f"discard_{_ur_id}", use_container_width=True):
                        try:
                            db_discard_run(_ur_id)
                            logger.info("User discarded unfinished run %s", _ur_id)
                        except Exception:
                            logger.exception("Failed to discard run %s", _ur_id)
                        if ss.get("resume_run_id") == _ur_id:
                            ss.resume_run_id = ""
                        st.rerun()
        if ss.get("resume_run_id"):
            st.info(f"Ready to resume run `{ss.resume_run_id}` — upload the original Excel file below to continue.")

    # Step indicators at the top
    has_file = ss.get("file_bytes") is not None or ss.completed
    s1_done = has_file
    s3_active = ss.running
    s3_done = ss.completed

    s1_cls = "done" if s1_done else "active"
    s2_cls = "done" if s1_done else ""
    s3_cls = "done" if s3_done else ("active" if s3_active else "")

    st.markdown(f'''
    <div style="display:flex; align-items:flex-start; gap:0; margin-bottom:1.5rem; position:relative;">
        <div class="step-card {s1_cls}" style="flex:1; animation: fadeInUp 0.4s ease-out;">
            <span class="step-num {'done' if s1_done else ''}">{'&#10003;' if s1_done else '1'}</span>
            <span class="step-icon">&#128196;</span>
            <div class="step-title">Upload File</div>
            <div class="step-desc">Excel file with stock numbers</div>
        </div>
        <div style="display:flex;align-items:center;padding-top:2.5rem;color:{_muted};font-size:1.5rem;margin:0 -0.3rem;">&#10132;</div>
        <div class="step-card {s2_cls}" style="flex:1; animation: fadeInUp 0.5s ease-out;">
            <span class="step-num {'done' if s1_done else ''}">{'&#10003;' if s1_done else '2'}</span>
            <span class="step-icon">&#9889;</span>
            <div class="step-title">Choose Speed</div>
            <div class="step-desc">Select search speed</div>
        </div>
        <div style="display:flex;align-items:center;padding-top:2.5rem;color:{_muted};font-size:1.5rem;margin:0 -0.3rem;">&#10132;</div>
        <div class="step-card {s3_cls}" style="flex:1; animation: fadeInUp 0.6s ease-out;">
            <span class="step-num {'done' if s3_done else ''}">{'&#10003;' if s3_done else '3'}</span>
            <span class="step-icon">&#128640;</span>
            <div class="step-title">Start & Download</div>
            <div class="step-desc">Run search & get results</div>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    # ── Step 1: Upload & Config ──
    st.markdown('<div class="sec">Step 1 - Upload &amp; Configure</div>', unsafe_allow_html=True)

    cu, cf, cn = st.columns([2, 1, 1])
    with cu:
        target_url = st.text_input("Target Website URL", value=DEFAULT_URL)
    with cf:
        uploaded_file = st.file_uploader("Upload Excel", type=["xlsx", "xls"])
    with cn:
        custom_name = st.text_input("Save File As:", value="3LINES_Results")
        ss.custom_name = custom_name

    with st.expander("Advanced Settings (Priority & Blacklist)", expanded=False):
        f1, f2 = st.columns(2)
        with f1:
            priority_input = st.text_input("Priority Companies", value="", placeholder="e.g. AMETEK, SAMI, BOEING")
        with f2:
            blacklist_input = st.text_input("Blacklisted Companies", value="", placeholder="e.g. HARSCO, ACME")
            st.markdown('<p class="blwarn">Blacklist adds extra processing time.</p>', unsafe_allow_html=True)
        skip_recent_chk = st.checkbox(
            "Skip stocks already successfully scraped within the last 24 hours",
            value=False,
            key="skip_recent_chk",
            help="OFF by default. When ON, stocks already in the local database from a successful run within the last 24h are skipped — useful when resuming a partial job.",
        )
        priority_targets = parse_comma_list(priority_input)
        blacklisted_companies = parse_comma_list(blacklist_input)

    if "priority_input" not in dir():
        priority_input = ""
    if "blacklist_input" not in dir():
        blacklist_input = ""
    priority_targets = parse_comma_list(priority_input)
    blacklisted_companies = parse_comma_list(blacklist_input)

    if uploaded_file:
        file_bytes = uploaded_file.getvalue(); ss.file_bytes = file_bytes
        detected_stocks, validation_error = load_stocks_strict(file_bytes)
        total_records = len(detected_stocks); ss.stock_count = total_records
        if validation_error and not ss.running and not ss.completed:
            st.error(validation_error)
        elif total_records == 0 and not ss.running and not ss.completed:
            st.error("File Rejected: Stock numbers must start from Row 2 in Column A")

        # Defensive: if batch_mode is stale from a previous batch group AND
        # the freshly uploaded file doesn't match the planned total, reset
        # the batch group. Prevents the batch UI showing stale plan against
        # the new file when the user uploads without clicking "Start a new file".
        if ss.batch_mode and ss.batch_plan and not ss.running and not ss.completed:
            _plan_total = sum(len(b["stocks"]) for b in ss.batch_plan)
            if _plan_total != total_records:
                logger.info(
                    "Resetting stale batch group: plan total=%d, new file=%d",
                    _plan_total, total_records,
                )
                ss.batch_mode = False
                ss.batch_plan = []
                ss.batch_current = 0
                ss.batch_run_ids = []
                ss.batch_group_id = ""
                ss["_batch_notice_dismissed"] = False

        # ── Large-file batch handling (Phase A notice + Phase B controls) ──
        # When the file exceeds BATCH_THRESHOLD rows, offer the user a
        # choice between a single run and splitting into batches of
        # BATCH_SIZE_DEFAULT. The empirical sweet spot at 3 bots is
        # ~1 h 22 m per 500-row batch.
        if total_records > BATCH_THRESHOLD and not ss.running and not ss.completed and not ss.batch_mode:
            _suggested_batches = (total_records + BATCH_SIZE_DEFAULT - 1) // BATCH_SIZE_DEFAULT
            st.info(
                f"**Large file detected — {total_records:,} rows.** "
                f"For stability, we recommend processing in **{_suggested_batches} batches of ~{BATCH_SIZE_DEFAULT} rows** "
                f"(the empirically-validated sweet spot at 3 bots, ~1 h 22 m per batch). "
                f"You can still continue with a single run, but expect roughly "
                f"{total_records // 60} minutes of wall-clock and higher risk of mid-run failures."
            )
            _bdec1, _bdec2, _ = st.columns([1, 1, 2])
            with _bdec1:
                if st.button(
                    f"Split into {_suggested_batches} batches",
                    key="batch_split_btn", use_container_width=True, type="primary",
                ):
                    base_name = (ss.get("custom_name", "") or "3LINES_Results").strip() or "3LINES_Results"
                    group_id = datetime.now().strftime("%Y%m%d_%H%M%S")
                    ss.batch_plan = _build_batch_plan(detected_stocks, BATCH_SIZE_DEFAULT, base_name, group_id)
                    ss.batch_current = 0
                    ss.batch_run_ids = []
                    ss.batch_group_id = group_id
                    ss.batch_mode = True
                    logger.info("Batch mode enabled: %d batches of %d, group=%s",
                                len(ss.batch_plan), BATCH_SIZE_DEFAULT, group_id)
                    st.rerun()
            with _bdec2:
                if st.button("Continue as single run", key="batch_single_btn", use_container_width=True):
                    # No-op functionally; clear the notice by setting a dismiss flag.
                    # Simplest path: set batch_mode False explicitly (already False)
                    # and hide the notice on next render via a session flag.
                    ss["_batch_notice_dismissed"] = True
                    st.rerun()
            if ss.get("_batch_notice_dismissed"):
                pass  # banner not re-shown after dismiss; reset on Run Again

        # ── Batch progress block (only when batch_mode is active) ──
        if ss.batch_mode and ss.batch_plan and not ss.running:
            _plan = ss.batch_plan
            _cur = ss.batch_current
            _total_batches = len(_plan)
            _total_stocks = sum(len(b["stocks"]) for b in _plan)
            _done_stocks = sum(len(b["stocks"]) for b in _plan[:_cur])
            with st.container():
                st.markdown(f"**Batch group:** `{ss.batch_group_id}` — {_total_batches} batches, {_total_stocks:,} stocks total")
                _lines = []
                for i, b in enumerate(_plan):
                    if i < _cur:
                        _lines.append(f"  ✓ Batch {b['index']} of {_total_batches}: {len(b['stocks'])} stocks — complete")
                    elif i == _cur:
                        _lines.append(f"  → Batch {b['index']} of {_total_batches}: {len(b['stocks'])} stocks — current")
                    else:
                        _lines.append(f"  ○ Batch {b['index']} of {_total_batches}: {len(b['stocks'])} stocks — pending")
                st.markdown("```\n" + "\n".join(_lines) + "\n```")
                st.caption(f"Overall progress: {_done_stocks:,} of {_total_stocks:,} stocks completed in this batch group.")
                _bcancel1, _bcancel2 = st.columns([1, 3])
                with _bcancel1:
                    if st.button("Cancel batch group", key="batch_cancel_btn", use_container_width=True):
                        ss.batch_mode = False
                        ss.batch_plan = []
                        ss.batch_current = 0
                        ss.batch_run_ids = []
                        ss.batch_group_id = ""
                        ss["_batch_notice_dismissed"] = False
                        logger.info("User cancelled batch group")
                        st.rerun()

        # ── Step 2: Speed ──
        st.markdown('<div class="sec">Step 2 - Choose Speed</div>', unsafe_allow_html=True)
        if "num_bots" not in ss:
            ss.num_bots = SMART_LIMIT
        if "speed_mode" not in ss:
            ss.speed_mode = "safe"
        # Each preset's bot count is capped at DH_MAX_BOTS so deployments can
        # tighten the ceiling via env var without code changes. "Maximum" maps
        # to whatever the configured cap is.
        _b_slow = min(1, DH_MAX_BOTS)
        _b_safe = min(3, DH_MAX_BOTS)
        _b_medium = min(6, DH_MAX_BOTS)
        _b_fast = DH_MAX_BOTS
        spm = {
            "slow":   {"b": _b_slow,   "l": "Careful",     "e": "\U0001f422",        "d": f"{_b_slow} bot{'s' if _b_slow>1 else ''} - safest"},
            "safe":   {"b": _b_safe,   "l": "Recommended", "e": "\U0001f6e1️",  "d": f"{_b_safe} bot{'s' if _b_safe>1 else ''} - stable"},
            "medium": {"b": _b_medium, "l": "Faster",      "e": "⚡",            "d": f"{_b_medium} bot{'s' if _b_medium>1 else ''} - quicker"},
            "fast":   {"b": _b_fast,   "l": "Maximum",     "e": "\U0001f680",        "d": f"{_b_fast} bot{'s' if _b_fast>1 else ''} - full power"},
        }
        # Defensive: re-derive num_bots from the current preset every render.
        # Streamlit's session_state persists ss.num_bots across reruns, so a
        # value cached when DH_MAX_BOTS was higher can leak forward (e.g.
        # ss.num_bots=3 after the operator lowers DH_MAX_BOTS to 2). The
        # active preset is the source of truth; num_bots is derived.
        ss.num_bots = spm.get(ss.speed_mode, spm["safe"])["b"]
        s1, s2, s3, s4 = st.columns(4)
        for col, mk in zip([s1, s2, s3, s4], ["slow", "safe", "medium", "fast"]):
            m = spm[mk]; sel = mk == ss.speed_mode
            best = " *" if mk == "safe" else ""
            check = "✅ " if sel else ""
            with col:
                if st.button(
                    f"{check}{m['e']} {m['l']}{best}\n{m['b']} bot{'s' if m['b']>1 else ''} - {m['d']}",
                    key=f"sp_{mk}", use_container_width=True,
                    type="primary" if sel else "secondary",
                ):
                    ss.speed_mode = mk; ss.num_bots = m["b"]; st.rerun()
        num_bots = ss.num_bots; sm = spm[ss.speed_mode]
        st.markdown(
            f'<div class="apbox"><span class="apt">{sm["e"]} {sm["l"]} Mode &mdash; {sm["b"]} bot{"s" if sm["b"]>1 else ""}</span><br>'
            f'<span class="apd">{sm["d"]}</span></div>',
            unsafe_allow_html=True,
        )

        # ── Step 3: Controls ──
        st.markdown('<div class="sec">Step 3 - Start Search</div>', unsafe_allow_html=True)
        c2, c3, c4 = st.columns([2, 1, 1])
        with c2:
            mx2 = max(total_records, 1)
            process_limit = st.number_input("How many to process (0 = all)", min_value=0, max_value=mx2, value=0, step=100, help="0 means process everything")
            st.caption(f"Will process all {total_records:,} records" if process_limit == 0 else f"Will process first {process_limit:,} of {total_records:,}")
        with c3:
            can = total_records > 0 and selenium_ok and not ss.running and not ss.completed and not validation_error
            # In batch mode, label the start button "Start Batch X of Y" so the
            # user knows which batch they're dispatching. Outside batch mode,
            # keep the original label.
            if ss.batch_mode and ss.batch_plan:
                _bx = ss.batch_current + 1
                _by = len(ss.batch_plan)
                _btn_label = f"Start Batch {_bx} of {_by}"
            else:
                _btn_label = "START SEARCH"
            start_btn = st.button(_btn_label, use_container_width=True, disabled=not can, type="primary")
            if not selenium_ok:
                st.caption("Selenium not installed")
        with c4:
            st.markdown('<div class="stop-btn-wrap">', unsafe_allow_html=True)
            if st.button("STOP", use_container_width=True, key="stop_m"):
                ss.stop_flag.set()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="hr2"></div>', unsafe_allow_html=True)

        # ── Live Tracker ──
        st.markdown('<div class="sec">Live Progress</div>', unsafe_allow_html=True)
        c1, c2, c3, c4, c5 = st.columns(5)
        m1_ph, m2_ph, m3_ph, m4_ph, m5_ph = c1.empty(), c2.empty(), c3.empty(), c4.empty(), c5.empty()
        cd2 = ss.processed; ct2 = ss.target if ss.target > 0 else total_records
        m1_ph.markdown(rmetric("Records", f"{cd2:,}/{ct2:,}", "g"), unsafe_allow_html=True)
        m2_ph.markdown(rmetric("Priority Stocks", f"{ss.priority_matches:,}", "b",
            help="Stocks where at least one supplier matched your priority list. Counted once per stock."),
            unsafe_allow_html=True)
        m3_ph.markdown(rmetric("Blacklisted Entries", f"{ss.blacklisted:,}", "r",
            help="Supplier entries excluded by your blacklist or built-in static filters (e.g. FINLAND, A486G). Counted per supplier entry, not per stock."),
            unsafe_allow_html=True)
        m4_ph.markdown(rmetric("Time Saved", f"{cd2*MINUTES_PER_ITEM_MANUAL:,}m", "p"), unsafe_allow_html=True)
        m5_ph.markdown(rmetric("ETA", "--", "b"), unsafe_allow_html=True)

        init_pct = min(ss.processed / ss.target, 1.0) if ss.target > 0 else 0
        progress_bar = st.progress(init_pct)
        status_ph = st.empty(); log_ph = st.empty(); stop_ph = st.empty()
        if not ss.running and not ss.completed:
            status_ph.markdown(f'<div class="sbox">{total_records:,} records ready to process</div>', unsafe_allow_html=True)
            log_ph.markdown(render_log([]), unsafe_allow_html=True)
        # auto_start_batch is set when the user clicks "Next batch" on a
        # completion screen; it triggers the next run_scraper call on
        # rerun without requiring a second Start click.
        _should_start = (start_btn or ss.get("auto_start_batch", False)) and not ss.running and not ss.completed
        if _should_start:
            ss.auto_start_batch = False
            try:
                if ss.batch_mode and ss.batch_plan:
                    # Dispatch the current batch: synthesize an Excel containing
                    # only the batch's stocks; override save_name; otherwise
                    # use the normal run_scraper code path unchanged.
                    _cur_batch = ss.batch_plan[ss.batch_current]
                    _batch_bytes = _make_batch_excel_bytes(_cur_batch["stocks"])
                    ss.custom_name = _cur_batch["save_name"]
                    run_scraper(
                        _batch_bytes, num_bots, 0, target_url, priority_targets, blacklisted_companies,
                        ss.stop_flag, status_ph, progress_bar, m1_ph, m2_ph, m3_ph, m4_ph, m5_ph, stop_ph, log_ph,
                        skip_recent=bool(skip_recent_chk),
                    )
                    # Capture the just-finished batch's run_id by looking it up
                    # via save_name (no orchestrator change needed).
                    rid = db_get_latest_run_id_for_save_name(_cur_batch["save_name"])
                    if rid and rid not in ss.batch_run_ids:
                        ss.batch_run_ids.append(rid)
                else:
                    run_scraper(
                        file_bytes, num_bots, process_limit, target_url, priority_targets, blacklisted_companies,
                        ss.stop_flag, status_ph, progress_bar, m1_ph, m2_ph, m3_ph, m4_ph, m5_ph, stop_ph, log_ph,
                        skip_recent=bool(skip_recent_chk),
                    )
                st.rerun()
            except Exception as e:
                ss.running = False
                st.error(f"Crashed: {e}")
                st.code(traceback.format_exc())

    elif not ss.completed:
        st.markdown(
            f'''<div class="upload-placeholder">
            <div class="up-icon">&#128196; &#10132; &#128194;</div>
            <div class="up-title">Upload Your Excel File</div>
            <div class="up-sub">
                Click <b>"Browse files"</b> above or drag & drop your file here<br>
                <span style="color:{_accent2}!important;">Supported: .xlsx and .xls</span> &mdash; Stock numbers in Column A, starting from Row 2
            </div>
        </div>''',
            unsafe_allow_html=True,
        )

    # ── Completion ──
    if ss.completed:
        elapsed = ss.elapsed; total = ss.target; ts5 = ss.processed * MINUTES_PER_ITEM_MANUAL
        em4, es4 = divmod(int(elapsed), 60); eh4, em4 = divmod(em4, 60)
        ed2 = f"{eh4}h {em4:02d}m {es4:02d}s" if eh4 else f"{em4}m {es4:02d}s"
        if not uploaded_file:
            st.markdown('<div class="sec">Final Results</div>', unsafe_allow_html=True)
            r1, r2, r3, r4, r5 = st.columns(5)
            r1.markdown(rmetric("Records", f"{ss.processed:,}/{total:,}", "g"), unsafe_allow_html=True)
            r2.markdown(rmetric("Priority Stocks", f"{ss.priority_matches:,}", "b",
                help="Stocks where at least one supplier matched your priority list. Counted once per stock."),
                unsafe_allow_html=True)
            r3.markdown(rmetric("Blacklisted Entries", f"{ss.blacklisted:,}", "r",
                help="Supplier entries excluded by your blacklist or built-in static filters (e.g. FINLAND, A486G). Counted per supplier entry, not per stock."),
                unsafe_allow_html=True)
            r4.markdown(rmetric("Time Saved", f"{ts5:,}m", "p"), unsafe_allow_html=True)
            r5.markdown(rmetric("Elapsed", ed2, "b"), unsafe_allow_html=True)
        if ss.final_log:
            st.markdown(render_log(ss.final_log), unsafe_allow_html=True)
        if ss.perf_data and len(ss.perf_data) > 2:
            with st.expander("Performance Chart", expanded=False):
                pdf = pd.DataFrame(ss.perf_data)
                pdf["rpm"] = pdf["records"].diff().fillna(0) * 60
                if "elapsed" in pdf.columns:
                    ed3 = pdf["elapsed"].diff().fillna(1).replace(0, 1)
                    pdf["rpm"] = (pdf["records"].diff().fillna(0) / ed3 * 60)
                pdf["rpm"] = pdf["rpm"].clip(lower=0)
                st.line_chart(pdf[["elapsed", "rpm"]].rename(columns={"elapsed": "Time(s)", "rpm": "Rec/min"}).set_index("Time(s)"))
        if ss.output_bytes:
            if ss.stopped:
                st.balloons()
                rem2 = total - ss.processed if total > ss.processed else 0
                st.markdown(
                    f'<div class="sbanner"><div class="st2">Stopped &amp; Saved</div>'
                    f'<div class="sm">{ss.processed:,} rows &bull; {rem2:,} remaining &bull; {ss.priority_matches:,} priority &bull; {ed2}</div></div>',
                    unsafe_allow_html=True,
                )
            else:
                st.balloons()
                st.markdown(
                    f'<div class="dbanner"><div class="dt">Search Complete!</div>'
                    f'<div class="dm">{ss.processed:,} rows &bull; {ss.priority_matches:,} priority &bull; {ss.blacklisted:,} blacklisted &bull; {ss.errors:,} errors</div></div>',
                    unsafe_allow_html=True,
                )
            if not ss.auto_downloaded:
                b64 = base64.b64encode(ss.output_bytes).decode()
                components.html(
                    f'<script>var a=document.createElement("a");a.href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}";a.download="{ss.output_name}";document.body.appendChild(a);a.click();</script>',
                    height=0,
                )
                ss.auto_downloaded = True
            if ss.final_data:
                with st.expander("Preview Results (first 10 rows)", expanded=False):
                    pf = pd.DataFrame(ss.final_data[:10])
                    cs2 = ["Stock Number"] + [c for c in pf.columns if c != "Stock Number"]
                    st.dataframe(pf[cs2], use_container_width=True)
            st.markdown('<div class="sec">Download Your Files</div>', unsafe_allow_html=True)
            d1, d2, d3 = st.columns(3)
            with d1:
                st.download_button(
                    "Download Excel",
                    data=ss.output_bytes,
                    file_name=ss.output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with d2:
                cv = build_csv(ss.final_data, priority_targets, blacklisted_companies)
                if cv:
                    st.download_button("Download CSV", data=cv, file_name=ss.output_name.replace(".xlsx", ".csv"), mime="text/csv", use_container_width=True)
            with d3:
                jv = build_json(ss.final_data, priority_targets, blacklisted_companies)
                if jv:
                    st.download_button("Download JSON", data=jv, file_name=ss.output_name.replace(".xlsx", ".json"), mime="application/json", use_container_width=True)
            # Failed-stocks export — only shown when at least one stock failed
            # to produce a result after the retry pass.
            if ss.failed_stocks:
                _failed_xlsx = build_failed_excel(ss.failed_stocks)
                if _failed_xlsx:
                    st.download_button(
                        f"Download Failed Stocks ({len(ss.failed_stocks):,})",
                        data=_failed_xlsx,
                        file_name=ss.output_name.replace(".xlsx", "_FAILED.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Stocks that did not produce a result, even after the auto-retry pass. Re-feed this file into the Scraper tab to try again.",
                    )
        elif ss.smart_skipped:
            st.info("All stocks were scraped recently — nothing to do. Disable smart-skip to re-run them.")
        else:
            st.error(f"No results found. Errors: {ss.errors:,} | Time: {int(elapsed)}s")

        # ── Phase B: batch-aware action buttons ──
        # If we're in batch mode and more batches remain, show "Next batch" +
        # "Stop batch group". On the final batch, surface the combined export
        # buttons in addition to the per-batch downloads above.
        if ss.batch_mode and ss.batch_plan:
            _bx = ss.batch_current + 1
            _by = len(ss.batch_plan)
            _has_more = _bx < _by

            if _has_more:
                # More batches to run — show Next batch + Stop
                _nb1, _nb2, _nb3 = st.columns([2, 1, 1])
                with _nb1:
                    if st.button(
                        f"Next batch ({_bx + 1} of {_by})",
                        key=f"next_batch_{ss.batch_current}",
                        use_container_width=True, type="primary",
                    ):
                        logger.info("User clicked Next batch: advancing from %d to %d (group=%s)",
                                    ss.batch_current, ss.batch_current + 1, ss.batch_group_id)
                        ss.batch_current += 1
                        _reset_per_run_state(defaults)
                        ss.auto_start_batch = True  # triggers run on rerun
                        st.rerun()
                with _nb2:
                    if st.button("Stop batch group", key="batch_stop_group", use_container_width=True):
                        logger.info("User stopped batch group after batch %d/%d", _bx, _by)
                        ss.batch_mode = False
                        ss.batch_plan = []
                        ss.batch_current = 0
                        ss.batch_group_id = ""
                        ss["_batch_notice_dismissed"] = False
                        st.rerun()
            else:
                # Final batch completed — surface combined exports
                st.markdown('<div class="sec">Combined Batch Group Exports</div>', unsafe_allow_html=True)
                st.success(
                    f"All {_by} batches complete. Combined exports below aggregate results "
                    f"from {len(ss.batch_run_ids)} batch run(s)."
                )
                try:
                    _combined_data = db_get_results_for_run_ids(ss.batch_run_ids)
                    _combined_failed = db_get_failed_stocks_for_run_ids(ss.batch_run_ids)
                except Exception:
                    logger.exception("Failed to load combined batch results")
                    _combined_data = []
                    _combined_failed = []

                _cb1, _cb2, _cb3 = st.columns(3)
                _group_label = f"{(ss.get('custom_name','').split('_batch_')[0] or 'batch_group')}_combined_{ss.batch_group_id}"
                with _cb1:
                    _cxlsx, _, _, _ = build_excel(_combined_data, priority_targets, blacklisted_companies)
                    if _cxlsx:
                        st.download_button(
                            f"Download Combined Excel ({len(_combined_data):,} rows)",
                            data=_cxlsx,
                            file_name=f"{_group_label}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                with _cb2:
                    _ccsv = build_csv(_combined_data, priority_targets, blacklisted_companies)
                    if _ccsv:
                        st.download_button(
                            f"Download Combined CSV ({len(_combined_data):,} rows)",
                            data=_ccsv,
                            file_name=f"{_group_label}.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )
                with _cb3:
                    _cjson = build_json(_combined_data, priority_targets, blacklisted_companies)
                    if _cjson:
                        st.download_button(
                            f"Download Combined JSON ({len(_combined_data):,} rows)",
                            data=_cjson,
                            file_name=f"{_group_label}.json",
                            mime="application/json",
                            use_container_width=True,
                        )

                if _combined_failed:
                    _cfx = build_failed_excel(_combined_failed)
                    if _cfx:
                        st.download_button(
                            f"Download All Failed Stocks ({len(_combined_failed):,})",
                            data=_cfx,
                            file_name=f"{_group_label}_ALL_FAILED.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            help="Stocks that did not produce a result across all batches in this group.",
                        )

                if st.button("Start a new file", key="batch_start_new", use_container_width=True, type="primary"):
                    # Full reset including batch state
                    for k, v in defaults.items():
                        ss[k] = v
                    ss.stop_flag.clear()
                    st.rerun()

        # Non-batch (single-run) flow gets the standard Run Again button
        if not ss.batch_mode:
            if st.button("Run Again", use_container_width=True, type="primary"):
                for k, v in defaults.items():
                    ss[k] = v
                ss.stop_flag.clear(); st.rerun()
    if ss.run_history:
        with st.expander(f"Run History ({len(ss.run_history)} runs)", expanded=False):
            for i, h in enumerate(reversed(ss.run_history)):
                ic = "Stopped" if h.get("stopped") else "Done"
                st.markdown(
                    f'<div class="hrow"><span>#{len(ss.run_history)-i} ({ic})</span>'
                    f'<span class="hd">{h["date"]}</span><span class="hr2c">{h["records"]:,}/{h["total"]:,}</span>'
                    f'<span class="hp">{h["priority"]:,} priority</span><span class="ht">{h["elapsed"]}</span></div>',
                    unsafe_allow_html=True,
                )
            if st.button("Clear History"):
                ss.run_history = []
                st.rerun()
