"""
3LINES DataHunter v14.0 - Premium Edition
Dynamic hardware inspection: Available RAM + Live CPU Load via psutil.
Safe Bots = Available_RAM / 0.6 GB per bot, halved if CPU > 70%.
Features: Run History, Multi-Format Export, Performance Chart,
Auto-Retry, Dark/Light Theme Toggle, Data Preview.
Strict Column A validation from Row 2. Dual filtering preserved.
Premium UI with tabbed navigation and professional design.
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import re
import io
import os
import time
import json
import base64
import shutil
import sqlite3
import threading
from datetime import datetime
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SQLITE DATABASE LAYER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "datahunter_local.db")


def _get_db():
    """Get a SQLite connection (one per thread)."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_local_db():
    """Create tables if they don't exist."""
    conn = _get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS runs (
            run_id          TEXT PRIMARY KEY,
            created_at      TEXT,
            status          TEXT DEFAULT 'completed',
            save_name       TEXT,
            total_stocks    INTEGER DEFAULT 0,
            processed       INTEGER DEFAULT 0,
            priority_count  INTEGER DEFAULT 0,
            blacklisted     INTEGER DEFAULT 0,
            errors          INTEGER DEFAULT 0,
            elapsed         TEXT,
            was_stopped     INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS run_results (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id      TEXT REFERENCES runs(run_id),
            stock_number TEXT,
            result_data TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_run_results ON run_results(run_id);
    """)
    conn.commit()
    conn.close()


def db_save_run(run_id, save_name, total_stocks, processed,
                priority_count, blacklisted, errors, elapsed, was_stopped, results):
    conn = _get_db()
    conn.execute(
        """INSERT OR REPLACE INTO runs
           (run_id, created_at, status, save_name, total_stocks, processed,
            priority_count, blacklisted, errors, elapsed, was_stopped)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (run_id, datetime.now().isoformat(), "completed",
         save_name, total_stocks, processed,
         priority_count, blacklisted, errors, elapsed, int(was_stopped)),
    )
    if results:
        conn.execute("DELETE FROM run_results WHERE run_id=?", (run_id,))
        conn.executemany(
            "INSERT INTO run_results (run_id, stock_number, result_data) VALUES (?, ?, ?)",
            [(run_id, r.get("Stock Number", ""), json.dumps(r, ensure_ascii=False))
             for r in results],
        )
    conn.commit()
    conn.close()


def db_get_all_runs():
    conn = _get_db()
    rows = conn.execute("SELECT * FROM runs ORDER BY created_at DESC").fetchall()
    result = [dict(r) for r in rows]
    conn.close()
    return result


def db_get_all_results():
    conn = _get_db()
    rows = conn.execute("""
        SELECT rr.stock_number, rr.result_data, r.run_id, r.created_at, r.save_name
        FROM run_results rr
        JOIN runs r ON rr.run_id = r.run_id
        ORDER BY r.created_at DESC, rr.id ASC
    """).fetchall()
    result = []
    for row in rows:
        data = json.loads(row["result_data"])
        data["_run_id"] = row["run_id"]
        data["_date"] = row["created_at"][:10]
        data["_save_name"] = row["save_name"]
        result.append(data)
    conn.close()
    return result


def db_get_run_results(run_id):
    conn = _get_db()
    rows = conn.execute(
        "SELECT result_data FROM run_results WHERE run_id=? ORDER BY id",
        (run_id,)
    ).fetchall()
    result = [json.loads(r["result_data"]) for r in rows]
    conn.close()
    return result


def db_get_total_stats():
    conn = _get_db()
    row = conn.execute("""
        SELECT COUNT(*) as total_runs,
               COALESCE(SUM(processed), 0) as total_processed,
               COALESCE(SUM(priority_count), 0) as total_priority,
               COALESCE(SUM(blacklisted), 0) as total_blacklisted,
               COALESCE(SUM(errors), 0) as total_errors,
               (SELECT COUNT(*) FROM run_results) as total_records
        FROM runs
    """).fetchone()
    result = dict(row)
    conn.close()
    return result


def db_delete_run(run_id):
    conn = _get_db()
    conn.execute("DELETE FROM run_results WHERE run_id=?", (run_id,))
    conn.execute("DELETE FROM runs WHERE run_id=?", (run_id,))
    conn.commit()
    conn.close()


def db_clear_all():
    conn = _get_db()
    conn.execute("DELETE FROM run_results")
    conn.execute("DELETE FROM runs")
    conn.commit()
    conn.close()


init_local_db()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  REAL-TIME RESOURCE AUDITING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PSUTIL_OK = False
try:
    import psutil
    PSUTIL_OK = True
except ImportError:
    pass


def get_system_status():
    if PSUTIL_OK:
        mem = psutil.virtual_memory()
        available_gb = round(mem.available / (1024 ** 3), 1)
        total_gb = round(mem.total / (1024 ** 3), 1)
        cpu_load = psutil.cpu_percent(interval=1)
        cpu_cores = psutil.cpu_count(logical=True) or os.cpu_count() or 2
    else:
        available_gb = 4.0
        total_gb = 4.0
        cpu_load = 0.0
        cpu_cores = os.cpu_count() or 2

    safe_bots = int(available_gb / 0.6)
    if cpu_load > 70:
        safe_bots = max(1, safe_bots // 2)
    safe_bots = max(1, safe_bots)

    return {
        "available_gb": available_gb,
        "total_gb": total_gb,
        "cpu_load": cpu_load,
        "cpu_cores": cpu_cores,
        "safe_bots": safe_bots,
    }


_SYS = get_system_status()
AVAILABLE_GB = _SYS["available_gb"]
TOTAL_GB = _SYS["total_gb"]
CPU_LOAD = _SYS["cpu_load"]
CPU_CORES = _SYS["cpu_cores"]
SMART_LIMIT = _SYS["safe_bots"]

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

# ── Page Config ──
st.set_page_config(
    page_title="3LINES DataHunter", page_icon="3L",
    layout="wide", initial_sidebar_state="collapsed"
)

# ── Session State Defaults ──
defaults = dict(
    running=False, completed=False, stopped=False,
    output_bytes=None, output_name="",
    processed=0, target=0, priority_matches=0, blacklisted=0, errors=0,
    start_time=0.0, elapsed=0.0,
    autosave_bytes=None, autosave_name="",
    stock_count=0, file_bytes=None, final_log=[],
    perf_data=[],
    failed_stocks=[],
    final_data=[],
    auto_downloaded=False,
)
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if "stop_flag" not in st.session_state:
    st.session_state.stop_flag = threading.Event()

if "run_history" not in st.session_state:
    st.session_state.run_history = []

if "theme" not in st.session_state:
    st.session_state.theme = "dark"

if "active_tab" not in st.session_state:
    st.session_state.active_tab = "scraper"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PREMIUM CSS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
is_dark = st.session_state.theme == "dark"

if is_dark:
    BG = "#060d1a"
    BG2 = "#0a1628"
    CARD = "#0f1d32"
    CARD2 = "#111d32"
    BORDER = "#1a2d4a"
    BORDER2 = "#243a5c"
    TEXT = "#e8ecf1"
    TEXT2 = "#c9d6e3"
    MUTED = "#6b7f99"
    INPUT_BG = "#0f1d32"
    ACCENT = "#3b82f6"
    ACCENT2 = "#60a5fa"
    GREEN = "#10b981"
    GREEN2 = "#34d399"
    RED = "#ef4444"
    RED2 = "#f87171"
    YELLOW = "#f59e0b"
    YELLOW2 = "#fbbf24"
    PURPLE = "#8b5cf6"
    PURPLE2 = "#a78bfa"
    HEADER_BG = "#002060"
    HEADER_BG2 = "#001845"
    GLOW_BLUE = "rgba(59,130,246,0.15)"
    GLOW_GREEN = "rgba(16,185,129,0.15)"
    GLOW_RED = "rgba(239,68,68,0.15)"
    GLOW_PURPLE = "rgba(139,92,246,0.15)"
    SHADOW = "rgba(0,0,0,0.4)"
else:
    BG = "#f0f4f8"
    BG2 = "#e8edf3"
    CARD = "#ffffff"
    CARD2 = "#f8fafc"
    BORDER = "#d1d9e6"
    BORDER2 = "#c0ccdb"
    TEXT = "#1a1a2e"
    TEXT2 = "#2d3748"
    MUTED = "#718096"
    INPUT_BG = "#ffffff"
    ACCENT = "#2563eb"
    ACCENT2 = "#3b82f6"
    GREEN = "#059669"
    GREEN2 = "#10b981"
    RED = "#dc2626"
    RED2 = "#ef4444"
    YELLOW = "#d97706"
    YELLOW2 = "#f59e0b"
    PURPLE = "#7c3aed"
    PURPLE2 = "#8b5cf6"
    HEADER_BG = "#002060"
    HEADER_BG2 = "#001845"
    GLOW_BLUE = "rgba(37,99,235,0.08)"
    GLOW_GREEN = "rgba(5,150,105,0.08)"
    GLOW_RED = "rgba(220,38,38,0.08)"
    GLOW_PURPLE = "rgba(124,58,237,0.08)"
    SHADOW = "rgba(0,0,0,0.08)"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600;700&display=swap');

/* ── Reset & Base ── */
.stApp {{
    background: {BG} !important;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}}
section[data-testid="stSidebar"] {{ display: none !important; }}
#MainMenu, footer, header {{ display: none !important; }}

.stApp, .stApp p, .stApp span, .stApp label, .stApp div,
.stApp li, .stApp h1, .stApp h2, .stApp h3, .stApp h4,
.stApp summary, .stApp td, .stApp th, .stApp a,
.stApp strong, .stApp em, .stApp code {{
    color: {TEXT} !important;
}}

/* ── Premium Header ── */
.premium-header {{
    background: linear-gradient(135deg, {HEADER_BG} 0%, {HEADER_BG2} 50%, #000d2b 100%);
    border-bottom: 1px solid rgba(96,165,250,0.2);
    padding: 0.9rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin: -1rem -1rem 0 -1rem;
    position: relative;
    overflow: hidden;
}}
.premium-header::before {{
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    bottom: 0;
    background: radial-gradient(ellipse at 20% 50%, rgba(59,130,246,0.1) 0%, transparent 70%);
    pointer-events: none;
}}
.header-brand {{
    display: flex;
    align-items: center;
    gap: 14px;
    position: relative;
    z-index: 1;
}}
.header-logo {{
    font-size: 1.7rem;
    font-weight: 900;
    color: #ffffff !important;
    letter-spacing: -0.5px;
    text-shadow: 0 0 20px rgba(96,165,250,0.3);
}}
.header-logo span {{
    color: {ACCENT2} !important;
    font-weight: 300;
}}
.header-divider {{
    width: 1px;
    height: 28px;
    background: rgba(255,255,255,0.15);
}}
.header-subtitle {{
    font-size: 0.78rem;
    color: rgba(192,210,235,0.8) !important;
    font-weight: 500;
    letter-spacing: 0.5px;
}}
.header-right {{
    display: flex;
    align-items: center;
    gap: 16px;
    position: relative;
    z-index: 1;
}}
.header-badge {{
    display: flex;
    align-items: center;
    gap: 6px;
    background: rgba(16,185,129,0.15);
    border: 1px solid rgba(16,185,129,0.3);
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.7rem;
    color: #34d399 !important;
    font-weight: 600;
}}
.header-badge .dot {{
    width: 6px;
    height: 6px;
    border-radius: 50%;
    background: #34d399;
    animation: pulse-dot 2s infinite;
}}
@keyframes pulse-dot {{
    0%, 100% {{ opacity: 1; }}
    50% {{ opacity: 0.3; }}
}}
.header-version {{
    font-size: 0.65rem;
    color: rgba(150,170,200,0.6) !important;
    font-weight: 500;
    font-family: 'JetBrains Mono', monospace;
}}

/* ── Tab Navigation ── */
.tab-nav {{
    background: {CARD};
    border-bottom: 1px solid {BORDER};
    display: flex;
    gap: 0;
    margin: 0 -1rem;
    padding: 0 1.5rem;
    box-shadow: 0 2px 8px {SHADOW};
}}
.tab-item {{
    padding: 0.85rem 1.5rem;
    font-size: 0.82rem;
    font-weight: 600;
    color: {MUTED} !important;
    cursor: pointer;
    border-bottom: 2px solid transparent;
    transition: all 0.2s ease;
    text-decoration: none !important;
    display: flex;
    align-items: center;
    gap: 8px;
    white-space: nowrap;
}}
.tab-item:hover {{
    color: {TEXT} !important;
    background: {GLOW_BLUE};
}}
.tab-item.active {{
    color: {ACCENT2} !important;
    border-bottom: 2px solid {ACCENT2};
    background: transparent;
}}
.tab-icon {{
    font-size: 1rem;
}}

/* ── Premium Cards ── */
.p-card {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 14px;
    padding: 1.4rem;
    box-shadow: 0 2px 12px {SHADOW};
    transition: all 0.25s ease;
}}
.p-card:hover {{
    border-color: {BORDER2};
    box-shadow: 0 4px 20px {SHADOW};
}}

/* ── Metric Cards ── */
.metric-card {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 14px;
    padding: 1.2rem 1rem;
    text-align: center;
    box-shadow: 0 2px 12px {SHADOW};
    transition: all 0.3s ease;
    position: relative;
    overflow: hidden;
}}
.metric-card::before {{
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 3px;
    border-radius: 14px 14px 0 0;
}}
.metric-card.mc-green::before {{ background: linear-gradient(90deg, {GREEN}, {GREEN2}); }}
.metric-card.mc-blue::before {{ background: linear-gradient(90deg, {ACCENT}, {ACCENT2}); }}
.metric-card.mc-red::before {{ background: linear-gradient(90deg, {RED}, {RED2}); }}
.metric-card.mc-purple::before {{ background: linear-gradient(90deg, {PURPLE}, {PURPLE2}); }}
.metric-card.mc-green {{ background: linear-gradient(180deg, {CARD} 60%, {'#0a1f14' if is_dark else '#f0fdf4'}); }}
.metric-card.mc-blue {{ background: linear-gradient(180deg, {CARD} 60%, {'#0a1525' if is_dark else '#eff6ff'}); }}
.metric-card.mc-red {{ background: linear-gradient(180deg, {CARD} 60%, {'#1f0a0a' if is_dark else '#fef2f2'}); }}
.metric-card.mc-purple {{ background: linear-gradient(180deg, {CARD} 60%, {'#150a25' if is_dark else '#f5f3ff'}); }}
.metric-card:hover {{
    transform: translateY(-2px);
    box-shadow: 0 8px 25px {SHADOW};
}}
.metric-card .m-value {{
    font-size: 1.85rem;
    font-weight: 800;
    line-height: 1.2;
    margin-top: 0.3rem;
    font-family: 'JetBrains Mono', monospace;
}}
.metric-card .m-label {{
    font-size: 0.65rem;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: {MUTED} !important;
    margin-top: 0.4rem;
    font-weight: 700;
}}
.m-green {{ color: {GREEN2} !important; }}
.m-blue  {{ color: {ACCENT2} !important; }}
.m-red   {{ color: {RED2} !important; }}
.m-purple{{ color: {PURPLE2} !important; }}

/* ── System Health Gauges ── */
.gauge-row {{
    display: flex;
    gap: 1rem;
    margin: 0.5rem 0;
}}
.gauge-item {{
    flex: 1;
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
}}
.gauge-ring {{
    width: 90px;
    height: 90px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    margin: 0 auto 0.5rem;
    position: relative;
}}
.gauge-ring .gauge-val {{
    font-size: 1.3rem;
    font-weight: 800;
    font-family: 'JetBrains Mono', monospace;
}}
.gauge-ring .gauge-unit {{
    font-size: 0.6rem;
    color: {MUTED} !important;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1px;
}}
.gauge-title {{
    font-size: 0.7rem;
    color: {MUTED} !important;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1.5px;
}}

/* ── Speed Mode Buttons ── */
.speed-grid {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.6rem;
    margin: 0.5rem 0;
}}
.speed-btn {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 10px;
    padding: 0.8rem 0.5rem;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s ease;
}}
.speed-btn:hover {{
    border-color: {ACCENT2};
    box-shadow: 0 0 15px {GLOW_BLUE};
}}
.speed-btn.active {{
    border-color: {ACCENT2};
    background: {'#0a1a30' if is_dark else '#eff6ff'};
    box-shadow: 0 0 20px {GLOW_BLUE};
}}
.speed-btn .sp-icon {{ font-size: 1.3rem; }}
.speed-btn .sp-name {{
    font-size: 0.75rem;
    font-weight: 700;
    color: {TEXT} !important;
    margin: 0.2rem 0;
}}
.speed-btn .sp-bots {{
    font-size: 0.65rem;
    color: {MUTED} !important;
    font-weight: 600;
    font-family: 'JetBrains Mono', monospace;
}}

/* ── Input Fields ── */
div[data-testid="stTextInput"] input,
div[data-testid="stNumberInput"] input,
div[data-testid="stTextArea"] textarea {{
    background: {INPUT_BG} !important;
    color: {TEXT} !important;
    -webkit-text-fill-color: {TEXT} !important;
    border: 1px solid {BORDER} !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    opacity: 1 !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
}}
div[data-testid="stTextInput"] input:focus,
div[data-testid="stNumberInput"] input:focus,
div[data-testid="stTextArea"] textarea:focus {{
    border-color: {ACCENT2} !important;
    box-shadow: 0 0 0 3px {GLOW_BLUE} !important;
}}
div[data-testid="stNumberInput"] button {{
    color: {TEXT} !important;
    background: {BORDER} !important;
    border: 1px solid {BORDER} !important;
}}

.stApp label,
div[data-testid="stWidgetLabel"] label,
div[data-testid="stWidgetLabel"] p {{
    color: {TEXT2} !important;
    font-weight: 700 !important;
    font-size: 0.85rem !important;
}}

.stApp .stCaption, .stApp small {{
    color: {MUTED} !important;
    font-weight: 500 !important;
}}

.stApp input::placeholder,
.stApp textarea::placeholder {{
    color: {MUTED} !important;
    opacity: 0.7 !important;
}}

/* ── Selectbox ── */
div[data-testid="stSelectbox"] > div > div {{
    background: {INPUT_BG} !important;
    color: {TEXT} !important;
    border: 1px solid {BORDER} !important;
    border-radius: 10px !important;
}}
div[data-testid="stSelectbox"] span {{ color: {TEXT} !important; }}
[data-baseweb="popover"] {{ background: {INPUT_BG} !important; border: 1px solid {BORDER} !important; border-radius: 10px !important; }}
[data-baseweb="popover"] ul {{ background: {INPUT_BG} !important; }}
[data-baseweb="popover"] li, [data-baseweb="menu"] li, ul[role="listbox"] li {{
    background: {INPUT_BG} !important; color: {TEXT} !important;
}}
[data-baseweb="popover"] li:hover, [data-baseweb="menu"] li:hover, ul[role="listbox"] li:hover {{
    background: {BORDER} !important;
}}
ul[role="listbox"] {{ background: {INPUT_BG} !important; }}

/* ── File Uploader ── */
div[data-testid="stFileUploader"] > div {{
    background: {INPUT_BG} !important;
    border: 2px dashed {BORDER2} !important;
    border-radius: 12px !important;
    transition: all 0.3s ease;
}}
div[data-testid="stFileUploader"] > div:hover {{
    border-color: {ACCENT2} !important;
    background: {GLOW_BLUE} !important;
}}
div[data-testid="stFileUploader"] span,
div[data-testid="stFileUploader"] small,
div[data-testid="stFileUploader"] p,
div[data-testid="stFileUploader"] div {{
    color: {TEXT2} !important;
}}
div[data-testid="stFileUploader"] button {{
    color: {TEXT} !important;
    background: {BORDER} !important;
    border: 1px solid {BORDER} !important;
}}

/* ── Expander ── */
div[data-testid="stExpander"] {{
    background: {CARD} !important;
    border: 1px solid {BORDER} !important;
    border-radius: 12px !important;
}}
div[data-testid="stExpander"] details summary {{
    color: {TEXT2} !important;
    font-weight: 700 !important;
}}

/* ── Alerts ── */
.stAlert, div[data-testid="stAlert"] {{
    background: {CARD} !important;
    color: {TEXT2} !important;
    border-color: {BORDER} !important;
    border-radius: 10px !important;
}}
.stAlert p, div[data-testid="stAlert"] p {{ color: {TEXT2} !important; }}

/* ── Progress Bar ── */
.stProgress > div > div > div > div {{
    background: linear-gradient(90deg, {ACCENT}, {GREEN2}) !important;
    border-radius: 8px;
    box-shadow: 0 0 10px {GLOW_BLUE};
}}
.stProgress > div > div > div {{
    background: {'#0f1d32' if is_dark else '#e2e8f0'} !important;
    border-radius: 8px;
}}

/* ── Dataframe ── */
.stDataFrame, div[data-testid="stDataFrame"] {{
    background: {CARD} !important;
    border-radius: 12px;
}}

/* ── All Buttons ── */
.stApp button {{
    color: {TEXT} !important;
    background: {CARD} !important;
    border: 1px solid {BORDER} !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    white-space: pre-line !important;
    line-height: 1.4 !important;
    padding: 0.5rem 0.8rem !important;
    transition: all 0.2s ease !important;
}}
.stApp button:hover {{
    background: {BORDER} !important;
    border-color: {BORDER2} !important;
}}
.stApp button p {{ color: inherit !important; white-space: pre-line !important; }}

/* Primary button (START) */
.stApp button[kind="primary"] {{
    background: linear-gradient(135deg, {GREEN}, #047857) !important;
    color: #ffffff !important;
    font-weight: 800 !important;
    font-size: 15px !important;
    border: none !important;
    height: 3.2em !important;
    box-shadow: 0 4px 15px {GLOW_GREEN} !important;
}}
.stApp button[kind="primary"]:hover {{
    background: linear-gradient(135deg, #047857, #065f46) !important;
    box-shadow: 0 6px 20px rgba(16,185,129,0.3) !important;
}}
.stApp button[kind="primary"] p {{ color: #ffffff !important; }}

/* Download buttons */
div[data-testid="stDownloadButton"] button {{
    background: linear-gradient(135deg, {ACCENT}, #1d4ed8) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    font-size: 0.85rem !important;
    border: none !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 15px {GLOW_BLUE} !important;
}}
div[data-testid="stDownloadButton"] button:hover {{
    background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
}}
div[data-testid="stDownloadButton"] button p {{ color: #ffffff !important; }}

/* ── Status Box ── */
.status-box {{
    padding: 14px 20px;
    border-radius: 10px;
    background: {CARD};
    border: 1px solid {BORDER};
    text-align: center;
    font-size: 0.9rem;
    font-weight: 700;
    color: {ACCENT2} !important;
    box-shadow: 0 2px 8px {SHADOW};
}}

/* ── Live Log ── */
.live-log {{
    background: {'#080e1a' if is_dark else '#f8fafc'};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 0.8rem 1.2rem;
    margin-top: 0.7rem;
    max-height: 220px;
    overflow-y: auto;
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.72rem;
    line-height: 1.7;
}}
.live-log .log-title {{
    color: {MUTED} !important;
    font-size: 0.6rem;
    text-transform: uppercase;
    letter-spacing: 2px;
    margin-bottom: 0.5rem;
    font-weight: 700;
}}
.live-log .log-line {{ color: {TEXT2} !important; }}
.live-log .log-line .bot-id {{ color: {ACCENT2} !important; font-weight: 700; }}
.live-log .log-line .stock-num {{ color: {GREEN2} !important; }}
.live-log .log-line .log-ok {{ color: {GREEN2} !important; }}
.live-log .log-line .log-err {{ color: {RED2} !important; font-weight: 700; }}
.live-log .log-line .log-priority {{ color: {YELLOW2} !important; font-weight: 700; }}
.live-log .log-line .log-blocked {{ color: {RED2} !important; font-weight: 700; }}

/* ── Banners ── */
.done-banner {{
    background: linear-gradient(135deg, {'#0a1f14' if is_dark else '#f0fff4'}, {'#0d2818' if is_dark else '#e6ffed'});
    border: 1px solid {'#166534' if is_dark else '#38a169'};
    border-radius: 14px;
    padding: 1.8rem;
    margin: 1rem 0;
    text-align: center;
    box-shadow: 0 4px 20px {GLOW_GREEN};
}}
.done-banner .done-title {{
    color: {GREEN2} !important;
    font-size: 1.3rem;
    font-weight: 800;
}}
.done-banner .done-meta {{
    color: {MUTED} !important;
    font-size: 0.8rem;
    margin-top: 0.5rem;
    font-family: 'JetBrains Mono', monospace;
}}

.stopped-banner {{
    background: linear-gradient(135deg, {'#1a1500' if is_dark else '#fffff0'}, {'#1f1a00' if is_dark else '#fefcbf'});
    border: 1px solid {'#854d0e' if is_dark else '#d69e2e'};
    border-radius: 14px;
    padding: 1.8rem;
    margin: 1rem 0;
    text-align: center;
    box-shadow: 0 4px 20px rgba(245,158,11,0.1);
}}
.stopped-banner .stopped-title {{
    color: {YELLOW2} !important;
    font-size: 1.3rem;
    font-weight: 800;
}}
.stopped-banner .stopped-meta {{
    color: {MUTED} !important;
    font-size: 0.8rem;
    margin-top: 0.5rem;
    font-family: 'JetBrains Mono', monospace;
}}

/* ── Autopilot Box ── */
.autopilot-box {{
    background: {'#0a1f14' if is_dark else '#f0fdf4'};
    border: 1px solid {'#166534' if is_dark else '#38a169'};
    border-radius: 10px;
    padding: 12px 16px;
    font-size: 0.82rem;
    font-weight: 600;
    color: {GREEN2} !important;
}}
.autopilot-box .ap-title {{
    font-weight: 800;
    font-size: 0.85rem;
    color: {GREEN2} !important;
}}
.autopilot-box .ap-detail {{
    color: {MUTED} !important;
    font-weight: 500;
    font-size: 0.75rem;
    font-family: 'JetBrains Mono', monospace;
}}

/* ── RAM Alert ── */
.ram-alert {{
    background: {'#1f0a0a' if is_dark else '#fef2f2'};
    border: 1px solid {RED};
    border-left: 4px solid {RED};
    border-radius: 10px;
    padding: 12px 16px;
    margin: 0.5rem 0;
    font-size: 0.82rem;
    font-weight: 700;
    color: {RED2} !important;
}}

/* ── Blacklist Warning ── */
.blacklist-warning {{
    color: {YELLOW2} !important;
    font-size: 0.78rem;
    font-weight: 600;
    margin: 0.3rem 0 0 0;
    padding: 0.4rem 0.7rem;
    background: {'#1a1500' if is_dark else '#fffff0'};
    border-left: 3px solid {YELLOW};
    border-radius: 0 6px 6px 0;
}}

/* ── Section Labels ── */
.section-label {{
    font-size: 0.8rem;
    font-weight: 800;
    color: {MUTED} !important;
    text-transform: uppercase;
    letter-spacing: 2.5px;
    margin: 1rem 0 0.5rem;
    display: flex;
    align-items: center;
    gap: 8px;
}}
.section-label::after {{
    content: '';
    flex: 1;
    height: 1px;
    background: {BORDER};
}}

/* ── Divider ── */
.hr {{
    height: 1px;
    background: linear-gradient(90deg, transparent, {BORDER}, transparent);
    margin: 1.2rem 0;
    border: none;
}}

/* ── History Row ── */
.history-row {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 10px;
    padding: 0.7rem 1.2rem;
    margin: 0.4rem 0;
    font-size: 0.8rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
    transition: all 0.2s ease;
}}
.history-row:hover {{
    border-color: {BORDER2};
    box-shadow: 0 2px 10px {SHADOW};
}}
.history-row span {{ color: {TEXT} !important; }}
.history-row .h-date {{ color: {MUTED} !important; font-weight: 500; font-family: 'JetBrains Mono', monospace; }}
.history-row .h-records {{ color: {GREEN2} !important; font-weight: 700; }}
.history-row .h-priority {{ color: {ACCENT2} !important; font-weight: 700; }}
.history-row .h-time {{ color: {PURPLE2} !important; font-weight: 600; font-family: 'JetBrains Mono', monospace; }}

/* ── Dashboard Stats Grid ── */
.dash-stats {{
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1rem 0;
}}
.dash-stat {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 12px;
    padding: 1.2rem;
    transition: all 0.2s ease;
}}
.dash-stat:hover {{
    border-color: {BORDER2};
    transform: translateY(-1px);
}}
.dash-stat .ds-label {{
    font-size: 0.65rem;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    color: {MUTED} !important;
    font-weight: 700;
    margin-bottom: 0.3rem;
}}
.dash-stat .ds-value {{
    font-size: 1.8rem;
    font-weight: 800;
    font-family: 'JetBrains Mono', monospace;
}}
.ds-green {{ color: {GREEN2} !important; }}
.ds-blue {{ color: {ACCENT2} !important; }}
.ds-red {{ color: {RED2} !important; }}
.ds-purple {{ color: {PURPLE2} !important; }}

/* ── System Info Bar ── */
.sys-bar {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
    gap: 0.8rem;
    margin: 0.8rem 0;
}}
.sys-item {{
    background: {CARD};
    border: 1px solid {BORDER};
    border-radius: 10px;
    padding: 0.8rem 1rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
}}
.sys-item .si-label {{
    font-size: 0.72rem;
    color: {MUTED} !important;
    font-weight: 600;
}}
.sys-item .si-value {{
    font-size: 0.85rem;
    font-weight: 700;
    font-family: 'JetBrains Mono', monospace;
}}

/* ── Scrollbar ── */
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: {BG}; }}
::-webkit-scrollbar-thumb {{ background: {BORDER2}; border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: {ACCENT}; }}

/* ── Footer ── */
.app-footer {{
    text-align: center;
    padding: 1.5rem 0;
    margin-top: 2rem;
    border-top: 1px solid {BORDER};
    font-size: 0.72rem;
    color: {MUTED} !important;
    font-weight: 500;
    letter-spacing: 0.5px;
}}
</style>
""", unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CONSTANTS & EXCEL STYLES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DEFAULT_URL = ""
STATIC_BLACKLIST = ["A486G", "FINLAND"]
AUTOSAVE_INTERVAL = 50
MINUTES_PER_ITEM_MANUAL = 2
MAX_LOG_LINES = 30
MAX_RETRIES = 2

H_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
H_FONT = Font(bold=True, size=11, color="FFFFFF")
PRIORITY_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
W_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BDR = Border(left=Side(style='thin'), right=Side(style='thin'),
             top=Side(style='thin'), bottom=Side(style='thin'))


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def load_stocks_strict(fb):
    xl = pd.ExcelFile(io.BytesIO(fb))
    df = pd.read_excel(io.BytesIO(fb), sheet_name=xl.sheet_names[0],
                       dtype=str, header=0)
    if df.empty or len(df.columns) == 0:
        return [], "\u274c File Rejected: Stock numbers must start from Row 2 in Column A"

    col_a = df.iloc[:, 0]
    if len(col_a) == 0 or pd.isna(col_a.iloc[0]) or str(col_a.iloc[0]).strip() == "":
        return [], "\u274c File Rejected: Stock numbers must start from Row 2 in Column A"

    stocks = []
    for val in col_a:
        if pd.notna(val) and str(val).strip():
            clean = str(val).strip().replace("-", "").replace(" ", "")
            if clean:
                stocks.append(clean)

    if not stocks:
        return [], "\u274c File Rejected: Stock numbers must start from Row 2 in Column A"
    return stocks, ""


def parse_comma_list(text):
    if not text or not text.strip():
        return []
    return [t.strip().upper() for t in text.split(",") if t.strip()]


def matches_company_list(mfg_name, company_list):
    if not mfg_name or not company_list:
        return False
    mu = mfg_name.strip().upper()
    for target in company_list:
        if target in mu:
            return True
    return False


def row_has_priority(row_dict, priority_list):
    if not priority_list:
        return False
    for key, val in row_dict.items():
        if key.startswith("MFG ") and val:
            if matches_company_list(str(val), priority_list):
                return True
    return False


def row_is_blacklisted(row_dict, blacklist):
    if not blacklist:
        return False
    for key, val in row_dict.items():
        if key.startswith("MFG ") and val and str(val).strip():
            if matches_company_list(str(val), blacklist):
                return True
    return False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CHROMIUM DETECTION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _find_binary():
    for p in ["/usr/bin/chromium", "/usr/bin/chromium-browser",
              "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable"]:
        if os.path.isfile(p):
            return p
    return shutil.which("chromium") or shutil.which("google-chrome")


def _find_driver():
    for p in ["/usr/bin/chromedriver", "/usr/lib/chromium/chromedriver",
              "/usr/lib/chromium-browser/chromedriver"]:
        if os.path.isfile(p):
            return p
    return shutil.which("chromedriver")


def make_driver():
    opts = Options()
    for flag in ["--headless=new", "--no-sandbox", "--disable-dev-shm-usage",
                 "--disable-gpu", "--disable-extensions", "--disable-notifications",
                 "--disable-popup-blocking", "--log-level=3", "--window-size=1200,800",
                 "--disable-software-rasterizer"]:
        opts.add_argument(flag)

    if os.name != "nt":
        opts.add_argument("--single-process")
        opts.add_argument("--no-zygote")

    opts.add_experimental_option('excludeSwitches', ['enable-logging'])
    opts.page_load_strategy = 'eager'

    chrome_bin = _find_binary()
    if chrome_bin:
        opts.binary_location = chrome_bin

    driver_path = _find_driver()
    if driver_path:
        drv = webdriver.Chrome(service=Service(driver_path), options=opts)
    else:
        drv = webdriver.Chrome(options=opts)

    drv.set_page_load_timeout(45)
    drv.set_script_timeout(20)
    drv.implicitly_wait(8)
    return drv


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SCRAPER (DUAL FILTERING)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def scrape_one(drv, wt, stock, target_url, priority_targets, blacklisted_companies):
    s = stock.strip()
    try:
        try:
            box = wt.until(EC.presence_of_element_located((By.ID, "nALL")))
        except Exception:
            drv.get(target_url)
            time.sleep(3)
            box = wt.until(EC.presence_of_element_located((By.ID, "nALL")))

        box.clear()
        time.sleep(0.3)
        box.send_keys(s)
        time.sleep(0.2)
        box.send_keys(Keys.RETURN)
        time.sleep(2)

        src = drv.page_source
        if "Search Results:" in src or "results found" in src.lower():
            try:
                lks = (
                    drv.find_elements(By.XPATH,
                        "//a[contains(@href,'NIIN') or contains(@href,'niin')]") or
                    drv.find_elements(By.XPATH,
                        "//a[string-length(normalize-space(text()))=9 "
                        "and translate(text(),'0123456789','')='']") or
                    drv.find_elements(By.XPATH, "//table//tr//td//a")
                )
                for lk in (lks or []):
                    if lk.text.strip() and len(lk.text.strip()) >= 5:
                        lk.click()
                        time.sleep(2)
                        break
            except Exception:
                pass

        rows = drv.find_elements(By.TAG_NAME, "tr")
        fstock, niin = "", ""
        for r in rows:
            cells = r.find_elements(By.TAG_NAME, "td")
            if len(cells) >= 2:
                t = [c.text.strip() for c in cells]
                if t[0] == "NIIN:" and len(t) > 1:
                    niin = t[1]
                if t[0] == "FSC:" and len(t) > 1 and niin:
                    fstock = f"{t[1]}{niin}"

        time.sleep(1)
        rows = drv.find_elements(By.TAG_NAME, "tr")
        raw = []
        for r in rows:
            cells = r.find_elements(By.TAG_NAME, "td")
            if len(cells) < 3:
                continue
            t = [c.text.strip() for c in cells]
            fc = t[0].upper()
            if any(x in fc for x in ["NIIN", "FSC", "NSN", "MOE", "AAC", ":"]):
                continue
            if any(x in fc for x in STATIC_BLACKLIST):
                continue
            cage = -1
            for i, tx in enumerate(t):
                if tx and len(tx) == 5 and re.match(r'^[A-Z0-9]{5}$', tx):
                    cage = i
                    break
            if cage <= 0:
                continue
            pn = t[0].strip()
            if any(pn.upper().startswith(b) for b in
                   ["HUES", "ABGL", "SHPE", "FSC", "NIIN", "NSN", "MOE", "AAC",
                    "RNCC", "RNVC", "DAC", "RNAAC", "CAGE"]):
                continue
            if len(pn) <= 3:
                continue
            co = ""
            for j in range(cage + 1, min(cage + 4, len(t))):
                cd = t[j].strip()
                if len(cd) <= 5:
                    continue
                if cd.upper() in ["NATURAL", "BLACK", "RECTANGULAR", "MINIMUM"]:
                    continue
                if "INCH" in cd.upper():
                    continue
                co = cd
                break
            if pn and co and len(co) > 5:
                raw.append((pn, t[cage], co))

        if not fstock:
            fstock = s
        fstock = fstock.replace("-", "")

        priority_entries, other_entries, seen = [], [], set()
        blacklisted_count = 0

        for pn, _, co in raw:
            pn, co = pn.strip(), co.strip()
            if not pn or pn in seen:
                continue
            cu = co.upper()
            if any(w in cu for w in
                   ["HUES", "ABGL", "SHPE", "CRF,", "NATURAL",
                    "BLACK", "RECTANGULAR", "FSC", "NIIN"]):
                continue
            if len(co) <= 5:
                continue
            seen.add(pn)

            if matches_company_list(co, blacklisted_companies):
                blacklisted_count += 1
                continue

            if matches_company_list(co, priority_targets):
                priority_entries.append((pn, co))
            else:
                other_entries.append((pn, co))

        res = {"Stock Number": fstock}
        slot = 1
        for pn, mfg in priority_entries:
            res[f"P.NO {slot}"] = pn
            res[f"MFG {slot}"] = mfg
            slot += 1
        for pn, mfg in other_entries:
            res[f"P.NO {slot}"] = pn
            res[f"MFG {slot}"] = mfg
            slot += 1

        if slot == 1:
            res["P.NO 1"] = ""
            res["MFG 1"] = ""

        return res, "ok", blacklisted_count

    except Exception as e:
        em = str(e).lower()
        if "session" in em or "invalid session" in em:
            return None, "dead", 0
        return {"Stock Number": s, "P.NO 1": "", "MFG 1": ""}, "err", 0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  EXCEL BUILDER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_excel(data, priority_targets, blacklisted_companies):
    if not data:
        return None, 0, 0, 0

    clean_data = []
    excluded = 0
    for r in data:
        if row_is_blacklisted(r, blacklisted_companies):
            excluded += 1
        else:
            clean_data.append(r)

    if not clean_data:
        return None, 0, 0, excluded

    prio_rows = [r for r in clean_data if row_has_priority(r, priority_targets)]
    other_rows = [r for r in clean_data if not row_has_priority(r, priority_targets)]
    rows = prio_rows + other_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "3LINES Output"

    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    mx = max(
        (int(c.split()[-1]) for c in all_keys
         if c.startswith("P.NO ") or c.startswith("MFG ")),
        default=1
    )
    headers = ["Stock Number"]
    for i in range(1, mx + 1):
        headers += [f"P.NO {i}", f"MFG {i}"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.border = BDR
        cell.alignment = Alignment(horizontal='center', vertical='center')

    priority_count = 0
    for ri, rd in enumerate(rows, 2):
        is_prio = row_has_priority(rd, priority_targets)
        if is_prio:
            priority_count += 1
        for ci, h in enumerate(headers, 1):
            v = rd.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = BDR
            cell.font = Font(size=10)
            if is_prio:
                cell.fill = PRIORITY_FILL
            else:
                cell.fill = W_FILL
            if ci == 1:
                cell.number_format = '@'
                cell.value = str(v).strip() if v else ""

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
    ws.column_dimensions['A'].width = 20
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), priority_count, len(rows), excluded


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  CSV & JSON BUILDERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_csv(data, priority_targets, blacklisted_companies):
    if not data:
        return None
    clean_data = [r for r in data if not row_is_blacklisted(r, blacklisted_companies)]
    if not clean_data:
        return None
    prio_rows = [r for r in clean_data if row_has_priority(r, priority_targets)]
    other_rows = [r for r in clean_data if not row_has_priority(r, priority_targets)]
    rows = prio_rows + other_rows
    df = pd.DataFrame(rows)
    cols = ["Stock Number"] + [c for c in df.columns if c != "Stock Number"]
    df = df[cols]
    return df.to_csv(index=False).encode("utf-8")


def build_json(data, priority_targets, blacklisted_companies):
    if not data:
        return None
    clean_data = [r for r in data if not row_is_blacklisted(r, blacklisted_companies)]
    if not clean_data:
        return None
    prio_rows = [r for r in clean_data if row_has_priority(r, priority_targets)]
    other_rows = [r for r in clean_data if not row_has_priority(r, priority_targets)]
    rows = prio_rows + other_rows
    return json.dumps(rows, indent=2, ensure_ascii=False).encode("utf-8")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  UI RENDERERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def render_log_html(log_entries):
    lines = ""
    for e in log_entries:
        bot = e.get("bot", "?")
        stock = e.get("stock", "")
        status = e.get("status", "")
        num = e.get("num", "")
        if status == "start":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; Scraping <span class="stock-num">{stock}</span>...</div>')
        elif status == "ok":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; <span class="log-ok">Done</span></div>')
        elif status == "priority":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; <span class="log-priority">PRIORITY MATCH</span></div>')
        elif status == "blocked":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; <span class="log-blocked">BLACKLISTED</span></div>')
        elif status == "err":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; <span class="log-err">Error</span></div>')
        elif status == "dead":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'<span class="log-err">CRASHED: {stock} &#8212; restarting...</span></div>')
        elif status == "retry":
            lines += (f'<div class="log-line"><span class="bot-id">[Bot {bot}]</span> '
                      f'Record <span class="stock-num">#{num}</span> '
                      f'&#8594; <span class="log-priority">RETRY</span> '
                      f'<span class="stock-num">{stock}</span></div>')
    return f'<div class="live-log"><div class="log-title">Live Processing Log</div>{lines}</div>'


def render_metric(label, value, color_class="m-green"):
    mc_class = color_class.replace("m-", "mc-")
    return (f'<div class="metric-card {mc_class}">'
            f'<div class="m-value {color_class}">{value}</div>'
            f'<div class="m-label">{label}</div></div>')


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  ORCHESTRATOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def run(file_bytes, num_workers, limit, target_url,
        priority_targets, blacklisted_companies,
        stop_flag, status_ph, bar_ph, m1_ph, m2_ph, m3_ph, m4_ph, m5_ph,
        stop_ph, log_ph):
    ss = st.session_state
    ss.running = True
    ss.completed = False
    ss.stopped = False
    ss.processed = 0
    ss.priority_matches = 0
    ss.blacklisted = 0
    ss.errors = 0
    ss.output_bytes = None
    ss.output_name = ""
    ss.autosave_bytes = None
    ss.autosave_name = ""
    ss.perf_data = []
    ss.failed_stocks = []
    ss.final_data = []
    stop_flag.clear()

    stocks, err_msg = load_stocks_strict(file_bytes)
    if not stocks:
        status_ph.error(err_msg or "\u274c File Rejected: Stock numbers must start from Row 2 in Column A")
        ss.running = False
        return

    if limit > 0 and limit < len(stocks):
        stocks = stocks[:limit]

    total = len(stocks)
    ss.target = total
    t0 = time.time()
    ss.start_time = t0

    lock = threading.Lock()
    results = []
    ctr = {"done": 0, "priority": 0, "blacklisted": 0, "errors": 0}
    last_autosave = {"count": 0}
    log_entries = deque(maxlen=MAX_LOG_LINES)
    perf_points = []
    failed_list = []

    def do_autosave():
        with lock:
            snapshot = list(results)
        if not snapshot:
            return
        try:
            xb, pc, tr, ex = build_excel(snapshot, priority_targets, blacklisted_companies)
            if xb:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                ss.autosave_bytes = xb
                ss.autosave_name = f"3LINES_AutoSave_{ts}.xlsx"
        except Exception:
            pass

    def worker(wid, chunk, start_idx):
        drv = None
        restarts = 0

        def boot():
            nonlocal drv
            try:
                if drv:
                    drv.quit()
            except Exception:
                pass
            with lock:
                log_entries.append({"bot": wid, "stock": "", "status": "start", "num": "BOOT"})
            d = make_driver()
            w = WebDriverWait(d, 15)
            for attempt in range(3):
                try:
                    d.get(target_url)
                    time.sleep(3)
                    d.find_element(By.ID, "nALL")
                    with lock:
                        log_entries.append({"bot": wid, "stock": "", "status": "ok", "num": "READY"})
                    return d, w
                except Exception as page_err:
                    with lock:
                        log_entries.append({"bot": wid, "stock": str(page_err)[:40],
                                            "status": "err", "num": f"LOAD #{attempt+1}"})
                    if attempt < 2:
                        time.sleep(3)
            return d, w

        try:
            drv, wt = boot()
        except Exception as boot_err:
            with lock:
                ctr["done"] += len(chunk)
                ctr["errors"] += len(chunk)
                failed_list.extend(chunk)
                log_entries.append({"bot": wid, "stock": f"CHROME FAILED: {str(boot_err)[:50]}",
                                    "status": "dead", "num": f"ALL ({len(chunk)})"})
            return

        try:
            for ci, stk in enumerate(chunk):
                if stop_flag.is_set():
                    break
                record_num = start_idx + ci + 1
                with lock:
                    log_entries.append({"bot": wid, "stock": stk, "status": "start", "num": record_num})
                try:
                    res, status, bl_count = scrape_one(
                        drv, wt, stk, target_url, priority_targets, blacklisted_companies)

                    if status == "ok" and res and res.get("Stock Number", "").strip():
                        is_prio = row_has_priority(res, priority_targets)
                        with lock:
                            results.append(res)
                            ctr["done"] += 1
                            ctr["blacklisted"] += bl_count
                            if bl_count > 0:
                                log_entries.append({"bot": wid, "stock": stk,
                                                    "status": "blocked", "num": record_num})
                            if is_prio:
                                ctr["priority"] += 1
                                log_entries.append({"bot": wid, "stock": stk,
                                                    "status": "priority", "num": record_num})
                            else:
                                log_entries.append({"bot": wid, "stock": stk,
                                                    "status": "ok", "num": record_num})

                    elif status == "dead":
                        with lock:
                            log_entries.append({"bot": wid, "stock": stk,
                                                "status": "dead", "num": record_num})
                            failed_list.append(stk)
                        restarts += 1
                        if restarts > 10:
                            break
                        time.sleep(2)
                        try:
                            drv, wt = boot()
                        except Exception:
                            break

                    elif status == "err":
                        with lock:
                            if res:
                                results.append(res)
                            ctr["done"] += 1
                            ctr["errors"] += 1
                            failed_list.append(stk)
                            log_entries.append({"bot": wid, "stock": stk,
                                                "status": "err", "num": record_num})
                except Exception:
                    with lock:
                        ctr["done"] += 1
                        ctr["errors"] += 1
                        failed_list.append(stk)
                        log_entries.append({"bot": wid, "stock": stk,
                                            "status": "err", "num": record_num})
        except Exception:
            pass
        finally:
            try:
                drv.quit()
            except Exception:
                pass

    chunk_size = max(1, total // num_workers)
    chunks = []
    start_indices = []
    for i in range(num_workers):
        start = i * chunk_size
        end = start + chunk_size if i < num_workers - 1 else total
        if start < total:
            chunks.append(stocks[start:end])
            start_indices.append(start)

    with stop_ph:
        if st.button("STOP & SAVE", use_container_width=True, key="stop_btn"):
            stop_flag.set()

    pool = ThreadPoolExecutor(max_workers=len(chunks))
    futures = {
        pool.submit(worker, i + 1, ch, si): i + 1
        for i, (ch, si) in enumerate(zip(chunks, start_indices))
    }

    try:
        while any(not f.done() for f in futures):
            time.sleep(1)
            if stop_flag.is_set():
                break
            with lock:
                d = ctr["done"]
                p = ctr["priority"]
                bl = ctr["blacklisted"]
                e = ctr["errors"]
                log_snapshot = list(log_entries)

            ss.processed = d
            ss.priority_matches = p
            ss.blacklisted = bl
            ss.errors = e

            elapsed_now = time.time() - t0
            perf_points.append({"elapsed": round(elapsed_now, 1), "records": d})

            pct = min(d / total, 1.0) if total else 0
            bar_ph.progress(pct)
            status_ph.markdown(
                f'<div class="status-box">Processing Record <b>#{d:,}</b> of <b>{total:,}</b></div>',
                unsafe_allow_html=True)

            time_saved = d * MINUTES_PER_ITEM_MANUAL
            if d > 0:
                avg_per_record = elapsed_now / d
                remaining = total - d
                eta_seconds = int(remaining * avg_per_record)
                eta_m, eta_s = divmod(eta_seconds, 60)
                eta_h, eta_m = divmod(eta_m, 60)
                eta_str = f"{eta_h}h {eta_m:02d}m" if eta_h else f"{eta_m}m {eta_s:02d}s"
            else:
                eta_str = "Calculating..."
            m1_ph.markdown(render_metric("Records Completed", f"{d:,} / {total:,}", "m-green"), unsafe_allow_html=True)
            m2_ph.markdown(render_metric("Priority Matches", f"{p:,}", "m-blue"), unsafe_allow_html=True)
            m3_ph.markdown(render_metric("Blacklisted", f"{bl:,}", "m-red"), unsafe_allow_html=True)
            m4_ph.markdown(render_metric("Est. Time Saved", f"{time_saved:,} min", "m-purple"), unsafe_allow_html=True)
            m5_ph.markdown(render_metric("ETA Remaining", eta_str, "m-blue"), unsafe_allow_html=True)
            log_ph.markdown(render_log_html(log_snapshot), unsafe_allow_html=True)

            if d - last_autosave["count"] >= AUTOSAVE_INTERVAL:
                last_autosave["count"] = d
                do_autosave()

        was_stopped = stop_flag.is_set()
        if was_stopped:
            status_ph.markdown(
                '<div class="status-box">Stopping... collecting final data from all bots</div>',
                unsafe_allow_html=True)
            wait_start = time.time()
            while time.time() - wait_start < 15:
                if all(f.done() for f in futures):
                    break
                time.sleep(0.5)
                with lock:
                    d = ctr["done"]
                    log_snapshot = list(log_entries)
                ss.processed = d
                pct = min(d / total, 1.0) if total else 0
                bar_ph.progress(pct)
                status_ph.markdown(
                    f'<div class="status-box">Stopping... collecting data '
                    f'(<b>{d:,}</b> records saved so far)</div>',
                    unsafe_allow_html=True)

        for f in as_completed(futures):
            try:
                f.result(timeout=5)
            except Exception:
                pass

    except BaseException:
        stop_flag.set()
        raise

    # ── AUTO-RETRY failed records ──
    if not was_stopped:
        was_stopped = stop_flag.is_set()
    retry_stocks = []
    with lock:
        retry_stocks = list(set(failed_list))
    if retry_stocks and not was_stopped and len(retry_stocks) <= total * 0.5:
        status_ph.markdown(
            f'<div class="status-box">Auto-retrying <b>{len(retry_stocks)}</b> failed records...</div>',
            unsafe_allow_html=True)
        with lock:
            log_entries.append({"bot": "R", "stock": "", "status": "retry",
                                "num": f"{len(retry_stocks)} records"})
        try:
            retry_drv = make_driver()
            retry_wt = WebDriverWait(retry_drv, 15)
            retry_drv.get(target_url)
            time.sleep(3)
            for ri, stk in enumerate(retry_stocks):
                if stop_flag.is_set():
                    break
                with lock:
                    log_entries.append({"bot": "R", "stock": stk, "status": "retry",
                                        "num": f"R-{ri+1}"})
                try:
                    res, status, bl_count = scrape_one(
                        retry_drv, retry_wt, stk, target_url,
                        priority_targets, blacklisted_companies)
                    if status == "ok" and res and res.get("Stock Number", "").strip():
                        with lock:
                            results[:] = [r for r in results
                                          if r.get("Stock Number", "") != stk or
                                          (r.get("P.NO 1", "") != "")]
                            results.append(res)
                            ctr["blacklisted"] += bl_count
                            if row_has_priority(res, priority_targets):
                                ctr["priority"] += 1
                            ctr["errors"] = max(0, ctr["errors"] - 1)
                            log_entries.append({"bot": "R", "stock": stk,
                                                "status": "ok", "num": f"R-{ri+1}"})
                    else:
                        with lock:
                            log_entries.append({"bot": "R", "stock": stk,
                                                "status": "err", "num": f"R-{ri+1}"})
                except Exception:
                    pass
            retry_drv.quit()
        except Exception:
            pass

    # ── Finalize ──
    if was_stopped:
        status_ph.markdown(
            '<div class="status-box">Saving collected data...</div>',
            unsafe_allow_html=True)

    with lock:
        final = list(results)
        d = ctr["done"]
        p = ctr["priority"]
        bl = ctr["blacklisted"]
        e = ctr["errors"]
        log_snapshot = list(log_entries)

    elapsed = time.time() - t0
    ss.elapsed = elapsed
    ss.perf_data = perf_points
    ss.final_data = final

    if final:
        try:
            xb, pc, tr, ex = build_excel(final, priority_targets, blacklisted_companies)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = ss.get("custom_name", "3LINES_Results").strip() or "3LINES_Results"
            label = "Partial" if was_stopped else "Result"
            ss.output_bytes = xb
            ss.output_name = f"{base}_{label}_{ts}.xlsx"
            ss.processed = tr
            ss.priority_matches = pc
            ss.blacklisted = bl + ex
            ss.errors = e
        except Exception:
            ss.processed = d
            ss.priority_matches = p
            ss.blacklisted = bl
            ss.errors = e
    else:
        ss.processed = d
        ss.priority_matches = p
        ss.blacklisted = bl
        ss.errors = e

    ss.final_log = list(log_snapshot) if log_snapshot else []
    ss.running = False
    ss.completed = True
    ss.stopped = was_stopped
    pool.shutdown(wait=False)

    el_m, el_s = divmod(int(elapsed), 60)
    el_h, el_m = divmod(el_m, 60)
    elapsed_display = f"{el_h}h {el_m:02d}m {el_s:02d}s" if el_h else f"{el_m}m {el_s:02d}s"
    ss.run_history.append({
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "records": ss.processed,
        "total": total,
        "priority": ss.priority_matches,
        "blacklisted": ss.blacklisted,
        "errors": ss.errors,
        "elapsed": elapsed_display,
        "stopped": was_stopped,
    })

    try:
        run_id = f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        save_name = ss.get("custom_name", "3LINES_Results").strip() or "3LINES_Results"
        db_save_run(
            run_id=run_id,
            save_name=save_name,
            total_stocks=total,
            processed=ss.processed,
            priority_count=ss.priority_matches,
            blacklisted=ss.blacklisted,
            errors=ss.errors,
            elapsed=elapsed_display,
            was_stopped=was_stopped,
            results=final,
        )
    except Exception:
        pass

    pct = min(d / total, 1.0) if total else 0
    bar_ph.progress(pct)

    if was_stopped:
        status_ph.markdown('<div class="status-box">Stopped by user &#8212; data saved</div>',
                           unsafe_allow_html=True)
    else:
        status_ph.markdown('<div class="status-box">Processing complete</div>',
                           unsafe_allow_html=True)

    time_saved = d * MINUTES_PER_ITEM_MANUAL
    elapsed_final = int(time.time() - t0)
    ef_m, ef_s = divmod(elapsed_final, 60)
    ef_h, ef_m = divmod(ef_m, 60)
    elapsed_str = f"{ef_h}h {ef_m:02d}m {ef_s:02d}s" if ef_h else f"{ef_m}m {ef_s:02d}s"
    m1_ph.markdown(render_metric("Records Completed", f"{ss.processed:,} / {total:,}", "m-green"), unsafe_allow_html=True)
    m2_ph.markdown(render_metric("Priority Matches", f"{ss.priority_matches:,}", "m-blue"), unsafe_allow_html=True)
    m3_ph.markdown(render_metric("Blacklisted", f"{ss.blacklisted:,}", "m-red"), unsafe_allow_html=True)
    m4_ph.markdown(render_metric("Est. Time Saved", f"{time_saved:,} min", "m-purple"), unsafe_allow_html=True)
    m5_ph.markdown(render_metric("Total Elapsed", elapsed_str, "m-blue"), unsafe_allow_html=True)
    log_ph.markdown(render_log_html(log_snapshot), unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  PREMIUM UI
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ss = st.session_state

# ── Premium Header ──
ram_pct = round(((TOTAL_GB - AVAILABLE_GB) / TOTAL_GB) * 100) if TOTAL_GB > 0 else 0
st.markdown(f'''
<div class="premium-header">
    <div class="header-brand">
        <div class="header-logo">3LINES <span>DataHunter</span></div>
        <div class="header-divider"></div>
        <div class="header-subtitle">Real-Time Resource Auditing &mdash; Dual Filtering &mdash; Full Autopilot</div>
    </div>
    <div class="header-right">
        <div class="header-badge">
            <div class="dot"></div>
            RAM {AVAILABLE_GB}GB &bull; CPU {CPU_LOAD}% &bull; {SMART_LIMIT} Safe Bots
        </div>
        <div class="header-version">v14.0 PREMIUM</div>
    </div>
</div>
''', unsafe_allow_html=True)

# ── Tab Navigation ──
tab_map = {
    "dashboard": ("Dashboard", "grid"),
    "scraper": ("Scraper", "play-circle"),
    "database": ("Database", "database"),
    "settings": ("Settings", "settings"),
}

tab_icons = {"dashboard": "\u2588\u2588", "scraper": "\u25b6", "database": "\u26c1", "settings": "\u2699"}

active_tab = ss.active_tab

tab_html = '<div class="tab-nav">'
for key, (label, icon) in tab_map.items():
    cls = "tab-item active" if key == active_tab else "tab-item"
    tab_html += f'<div class="{cls}" id="tab-{key}">{tab_icons[key]} {label}</div>'
tab_html += '</div>'
st.markdown(tab_html, unsafe_allow_html=True)

# Streamlit tab buttons (hidden visual, functional)
tab_cols = st.columns(len(tab_map))
for i, key in enumerate(tab_map.keys()):
    with tab_cols[i]:
        if st.button(tab_map[key][0], key=f"nav_{key}", use_container_width=True,
                     type="primary" if key == active_tab else "secondary"):
            ss.active_tab = key
            st.rerun()

# Theme toggle
theme_col1, theme_col2 = st.columns([8, 1])
with theme_col2:
    theme_icon = "\u2600\ufe0f" if is_dark else "\U0001f319"
    theme_label = f"{theme_icon} {'Light' if is_dark else 'Dark'}"
    if st.button(theme_label, key="theme_toggle"):
        ss.theme = "light" if is_dark else "dark"
        st.rerun()

st.markdown('<div class="hr"></div>', unsafe_allow_html=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB: DASHBOARD
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if active_tab == "dashboard":
    stats = db_get_total_stats()

    # Stats cards
    st.markdown(f'''
    <div class="dash-stats">
        <div class="dash-stat">
            <div class="ds-label">Total Runs</div>
            <div class="ds-value ds-blue">{stats["total_runs"]}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Total Records</div>
            <div class="ds-value ds-green">{stats["total_records"]:,}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Priority Matches</div>
            <div class="ds-value ds-purple">{stats["total_priority"]:,}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Total Errors</div>
            <div class="ds-value ds-red">{stats["total_errors"]:,}</div>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    # System Health
    st.markdown('<div class="section-label">System Health</div>', unsafe_allow_html=True)

    cpu_color = RED2 if CPU_LOAD > 80 else (YELLOW2 if CPU_LOAD > 60 else GREEN2)
    ram_color = RED2 if ram_pct > 80 else (YELLOW2 if ram_pct > 60 else ACCENT2)

    st.markdown(f'''
    <div class="sys-bar">
        <div class="sys-item">
            <span class="si-label">CPU Load</span>
            <span class="si-value" style="color:{cpu_color} !important">{CPU_LOAD}%</span>
        </div>
        <div class="sys-item">
            <span class="si-label">RAM Used</span>
            <span class="si-value" style="color:{ram_color} !important">{ram_pct}%</span>
        </div>
        <div class="sys-item">
            <span class="si-label">Available RAM</span>
            <span class="si-value">{AVAILABLE_GB} GB</span>
        </div>
        <div class="sys-item">
            <span class="si-label">Total RAM</span>
            <span class="si-value">{TOTAL_GB} GB</span>
        </div>
        <div class="sys-item">
            <span class="si-label">CPU Cores</span>
            <span class="si-value">{CPU_CORES}</span>
        </div>
        <div class="sys-item">
            <span class="si-label">Safe Bots</span>
            <span class="si-value" style="color:{GREEN2} !important">{SMART_LIMIT}</span>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    # Recent Jobs
    st.markdown('<div class="section-label">Recent Jobs</div>', unsafe_allow_html=True)

    all_runs = db_get_all_runs()
    if all_runs:
        for run_item in all_runs[:10]:
            status_icon = "\u26a0\ufe0f" if run_item["was_stopped"] else "\u2705"
            st.markdown(
                f'<div class="history-row">'
                f'<span>{status_icon} {run_item["save_name"]}</span>'
                f'<span class="h-date">{run_item["created_at"][:16]}</span>'
                f'<span class="h-records">{run_item["processed"]:,} / {run_item["total_stocks"]:,}</span>'
                f'<span class="h-priority">{run_item["priority_count"]:,} priority</span>'
                f'<span class="h-time">{run_item["elapsed"]}</span>'
                f'</div>',
                unsafe_allow_html=True)
    else:
        st.info("No jobs yet. Go to the Scraper tab to start your first job.")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB: SCRAPER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif active_tab == "scraper":
    # ── Row 1: URL + File + Save Name ──
    col_url, col_file, col_name = st.columns([2, 1, 1])
    with col_url:
        target_url = st.text_input("Target Website URL", value=DEFAULT_URL,
                                   help="LogiQuest FLIS URL to scrape from.")
    with col_file:
        uploaded_file = st.file_uploader("Upload Excel", type=["xlsx", "xls"])
    with col_name:
        custom_name = st.text_input("Save File As:", value="3LINES_Results")
        ss.custom_name = custom_name

    # ── Advanced Settings ──
    with st.expander("Advanced Settings", expanded=False):
        f1, f2 = st.columns(2)
        with f1:
            priority_input = st.text_input(
                "Priority Companies (comma-separated)", value="",
                placeholder="e.g. AMETEK, SAMI, NASA, BOEING",
                help="Matching rows get GREEN highlight and sort to TOP.")
        with f2:
            blacklist_input = st.text_input(
                "Blacklisted Companies (comma-separated)", value="",
                placeholder="e.g. HARSCO, ACME CORP",
                help="Matching rows are COMPLETELY EXCLUDED from output.")
            st.markdown(
                '<p class="blacklist-warning">'
                '\u26a0\ufe0f Blacklist filtering adds extra processing time. '
                'Each record is cross-checked before inclusion.</p>',
                unsafe_allow_html=True)

        priority_targets = parse_comma_list(priority_input)
        blacklisted_companies = parse_comma_list(blacklist_input)
        parts = []
        if priority_targets:
            parts.append(f"Priority: {', '.join(priority_targets)}")
        if blacklisted_companies:
            parts.append(f"Blacklisted: {', '.join(blacklisted_companies)}")
        if parts:
            st.caption(" | ".join(parts))

    if "priority_input" not in dir():
        priority_input = ""
    if "blacklist_input" not in dir():
        blacklist_input = ""
    priority_targets = parse_comma_list(priority_input)
    blacklisted_companies = parse_comma_list(blacklist_input)

    # ── Main content ──
    if uploaded_file:
        file_bytes = uploaded_file.getvalue()
        ss.file_bytes = file_bytes

        detected_stocks, validation_error = load_stocks_strict(file_bytes)
        total_records = len(detected_stocks)
        ss.stock_count = total_records

        if validation_error and not ss.running and not ss.completed:
            st.error(validation_error)
        elif total_records == 0 and not ss.running and not ss.completed:
            st.error("\u274c File Rejected: Stock numbers must start from Row 2 in Column A")

        # ── Speed Mode ──
        st.markdown('<div class="section-label">Speed Mode</div>', unsafe_allow_html=True)

        if "num_bots" not in ss:
            ss.num_bots = SMART_LIMIT
        if "speed_mode" not in ss:
            ss.speed_mode = "safe"

        safe_bots = max(SMART_LIMIT, 1)
        medium_bots = min(safe_bots + 5, 15)
        if medium_bots <= safe_bots:
            medium_bots = safe_bots + 2

        speed_modes = {
            "slow":   {"bots": 1,           "label": "Slow",   "icon": "\U0001f422"},
            "safe":   {"bots": safe_bots,   "label": "Safe",   "icon": "\U0001f6e1\ufe0f"},
            "medium": {"bots": medium_bots, "label": "Medium", "icon": "\u26a1"},
            "fast":   {"bots": 20,          "label": "Fast",   "icon": "\U0001f680"},
        }

        sp1, sp2, sp3, sp4 = st.columns(4)
        ordered = ["slow", "safe", "medium", "fast"]
        for col, mode_key in zip([sp1, sp2, sp3, sp4], ordered):
            mode = speed_modes[mode_key]
            is_selected = mode_key == ss.speed_mode
            is_rec = mode_key == "safe"
            rec_tag = " (Best)" if is_rec else ""
            check = "\u2705 " if is_selected else ""
            with col:
                if st.button(
                    f"{check}{mode['icon']} {mode['label']}{rec_tag}\n{mode['bots']} bot{'s' if mode['bots'] > 1 else ''}",
                    key=f"speed_{mode_key}",
                    use_container_width=True,
                    type="primary" if is_selected else "secondary"):
                    ss.speed_mode = mode_key
                    ss.num_bots = mode["bots"]
                    st.rerun()

        num_bots = ss.num_bots
        sel_mode = speed_modes[ss.speed_mode]

        st.markdown(
            f'<div class="autopilot-box">'
            f'<span class="ap-title">\u2705 {sel_mode["icon"]} {sel_mode["label"]} '
            f'({sel_mode["bots"]} bot{"s" if sel_mode["bots"] > 1 else ""})</span>'
            f'<br><span class="ap-detail">'
            f'RAM: {AVAILABLE_GB} GB &nbsp;|&nbsp; '
            f'CPU: {CPU_LOAD}% &nbsp;|&nbsp; Safe Limit: {SMART_LIMIT}'
            f'</span>'
            f'</div>', unsafe_allow_html=True)

        if num_bots > SMART_LIMIT:
            st.markdown(
                f'<div class="ram-alert">'
                f'\U0001f6a8 <b>Warning!</b> Safe limit is <b>{SMART_LIMIT} bots</b> '
                f'based on {AVAILABLE_GB} GB free RAM. '
                f'Running <b>{num_bots} bots</b> may freeze your system.'
                f'</div>', unsafe_allow_html=True)

        # ── Process Limit / START / STOP ──
        st.markdown('<div class="section-label">Controls</div>', unsafe_allow_html=True)
        c2, c3, c4 = st.columns([2, 1, 1])
        with c2:
            max_val = max(total_records, 1)
            process_limit = st.number_input(
                "Process Limit", min_value=0, max_value=max_val, value=0, step=100,
                help="0 = process ALL records")
            if process_limit == 0:
                st.caption(f"All {total_records:,} records")
            else:
                st.caption(f"First {process_limit:,} of {total_records:,}")
        with c3:
            can_start = (total_records > 0 and SELENIUM_OK
                         and not ss.running and not ss.completed
                         and not validation_error)
            start_btn = st.button("START", use_container_width=True,
                                  disabled=not can_start, type="primary")
            if not SELENIUM_OK:
                st.caption("Selenium not installed")
        with c4:
            stop_btn = st.button("STOP", use_container_width=True, key="stop_main")
            if stop_btn:
                ss.stop_flag.set()

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        # ── Live Tracker ──
        st.markdown('<div class="section-label">Live Tracker</div>', unsafe_allow_html=True)
        m1, m2, m3, m4, m5 = st.columns(5)
        m1_ph = m1.empty()
        m2_ph = m2.empty()
        m3_ph = m3.empty()
        m4_ph = m4.empty()
        m5_ph = m5.empty()

        cur_done = ss.processed
        cur_total = ss.target if ss.target > 0 else total_records
        cur_priority = ss.priority_matches
        cur_bl = ss.blacklisted
        cur_time = cur_done * MINUTES_PER_ITEM_MANUAL

        m1_ph.markdown(render_metric("Records Completed", f"{cur_done:,} / {cur_total:,}", "m-green"), unsafe_allow_html=True)
        m2_ph.markdown(render_metric("Priority Matches", f"{cur_priority:,}", "m-blue"), unsafe_allow_html=True)
        m3_ph.markdown(render_metric("Blacklisted", f"{cur_bl:,}", "m-red"), unsafe_allow_html=True)
        m4_ph.markdown(render_metric("Est. Time Saved", f"{cur_time:,} min", "m-purple"), unsafe_allow_html=True)
        m5_ph.markdown(render_metric("ETA Remaining", "--", "m-blue"), unsafe_allow_html=True)

        # ── Progress + Status + Log ──
        init_pct = min(ss.processed / ss.target, 1.0) if ss.target > 0 else 0
        progress_bar = st.progress(init_pct)
        status_ph = st.empty()
        log_ph = st.empty()
        stop_ph = st.empty()

        if not ss.running and not ss.completed:
            status_ph.markdown(
                f'<div class="status-box">{total_records:,} search keys detected \u2014 Ready</div>',
                unsafe_allow_html=True)
            log_ph.markdown(render_log_html([]), unsafe_allow_html=True)

        # ── Start ──
        if start_btn and not ss.running and not ss.completed:
            try:
                run(file_bytes, num_bots, process_limit, target_url,
                    priority_targets, blacklisted_companies,
                    ss.stop_flag, status_ph, progress_bar,
                    m1_ph, m2_ph, m3_ph, m4_ph, m5_ph, stop_ph, log_ph)
                st.rerun()
            except Exception as e:
                ss.running = False
                st.error(f"**Scraping crashed:** {type(e).__name__}: {e}")
                import traceback
                st.code(traceback.format_exc())

    elif not ss.completed:
        st.markdown(f'''
        <div class="p-card" style="text-align:center; padding: 3rem;">
            <div style="font-size: 3rem; margin-bottom: 1rem;">&#128194;</div>
            <div style="font-size: 1.1rem; font-weight: 700; color: {TEXT} !important; margin-bottom: 0.5rem;">Upload an Excel File to Begin</div>
            <div style="font-size: 0.82rem; color: {MUTED} !important;">Drag & drop or click the upload button above. Accepts .xlsx and .xls files.</div>
        </div>
        ''', unsafe_allow_html=True)

    # ── Completion UI ──
    if ss.completed:
        elapsed = ss.elapsed
        time_saved = ss.processed * MINUTES_PER_ITEM_MANUAL
        total = ss.target

        el_m, el_s = divmod(int(elapsed), 60)
        el_h, el_m = divmod(el_m, 60)
        elapsed_display = f"{el_h}h {el_m:02d}m {el_s:02d}s" if el_h else f"{el_m}m {el_s:02d}s"

        if not uploaded_file:
            st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
            st.markdown('<div class="section-label">Final Results</div>', unsafe_allow_html=True)
            r1, r2, r3, r4, r5 = st.columns(5)
            r1.markdown(render_metric("Records Completed", f"{ss.processed:,} / {total:,}", "m-green"), unsafe_allow_html=True)
            r2.markdown(render_metric("Priority Matches", f"{ss.priority_matches:,}", "m-blue"), unsafe_allow_html=True)
            r3.markdown(render_metric("Blacklisted", f"{ss.blacklisted:,}", "m-red"), unsafe_allow_html=True)
            r4.markdown(render_metric("Est. Time Saved", f"{time_saved:,} min", "m-purple"), unsafe_allow_html=True)
            r5.markdown(render_metric("Total Elapsed", elapsed_display, "m-blue"), unsafe_allow_html=True)

        if ss.final_log:
            st.markdown(render_log_html(ss.final_log), unsafe_allow_html=True)

        # Performance Chart
        if ss.perf_data and len(ss.perf_data) > 2:
            with st.expander("Performance Chart", expanded=False):
                perf_df = pd.DataFrame(ss.perf_data)
                perf_df["records_per_min"] = perf_df["records"].diff().fillna(0) * 60
                if "elapsed" in perf_df.columns:
                    elapsed_diff = perf_df["elapsed"].diff().fillna(1).replace(0, 1)
                    perf_df["records_per_min"] = (perf_df["records"].diff().fillna(0) / elapsed_diff * 60)
                perf_df["records_per_min"] = perf_df["records_per_min"].clip(lower=0)
                chart_df = perf_df[["elapsed", "records_per_min"]].rename(
                    columns={"elapsed": "Time (sec)", "records_per_min": "Records / min"})
                st.line_chart(chart_df.set_index("Time (sec)"))

        if ss.output_bytes:
            if ss.stopped:
                st.balloons()
                remaining = total - ss.processed if total > ss.processed else 0
                st.markdown(
                    f'<div class="stopped-banner">'
                    f'<div class="stopped-title">Stopped &amp; Data Saved Successfully</div>'
                    f'<div class="stopped-meta">'
                    f'{ss.processed:,} rows &bull; '
                    f'{remaining:,} remaining &bull; '
                    f'{ss.priority_matches:,} priority &bull; '
                    f'{ss.blacklisted:,} blacklisted &bull; '
                    f'{ss.errors:,} errors &bull; '
                    f'{elapsed_display} elapsed</div></div>',
                    unsafe_allow_html=True)
            else:
                st.balloons()
                st.markdown(
                    f'<div class="done-banner">'
                    f'<div class="done-title">Scraping Complete</div>'
                    f'<div class="done-meta">{ss.processed:,} rows &bull; '
                    f'{ss.priority_matches:,} priority &bull; '
                    f'{ss.blacklisted:,} blacklisted &bull; '
                    f'{ss.errors:,} errors</div></div>',
                    unsafe_allow_html=True)

            # Auto-Download
            if not ss.auto_downloaded:
                b64 = base64.b64encode(ss.output_bytes).decode()
                components.html(f'''<script>
                var a = document.createElement('a');
                a.href = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}';
                a.download = '{ss.output_name}';
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                </script>''', height=0)
                ss.auto_downloaded = True

            # Data Preview
            if ss.final_data:
                with st.expander("Data Preview (first 10 rows)", expanded=False):
                    preview_df = pd.DataFrame(ss.final_data[:10])
                    cols = ["Stock Number"] + [c for c in preview_df.columns if c != "Stock Number"]
                    preview_df = preview_df[cols]
                    st.dataframe(preview_df, use_container_width=True)

            # Multi-Format Download
            st.markdown('<div class="section-label">Download Results</div>', unsafe_allow_html=True)
            dl1, dl2, dl3 = st.columns(3)
            with dl1:
                st.download_button(
                    label=f"Excel: {ss.output_name}",
                    data=ss.output_bytes,
                    file_name=ss.output_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            with dl2:
                csv_data = build_csv(ss.final_data, priority_targets, blacklisted_companies)
                if csv_data:
                    csv_name = ss.output_name.replace(".xlsx", ".csv")
                    st.download_button(
                        label=f"CSV: {csv_name}",
                        data=csv_data,
                        file_name=csv_name,
                        mime="text/csv",
                        use_container_width=True)
            with dl3:
                json_data = build_json(ss.final_data, priority_targets, blacklisted_companies)
                if json_data:
                    json_name = ss.output_name.replace(".xlsx", ".json")
                    st.download_button(
                        label=f"JSON: {json_name}",
                        data=json_data,
                        file_name=json_name,
                        mime="application/json",
                        use_container_width=True)
        else:
            st.error(
                f"**All {total:,} records attempted but 0 results scraped.**\n\n"
                f"Check: Is Chrome installed? Run `pip install selenium`.\n"
                f"Try with 1 bot first.\n\n"
                f"Errors: {ss.errors:,} | Elapsed: {int(elapsed)}s")

        if st.button("Run Again", use_container_width=True):
            for k, v in defaults.items():
                ss[k] = v
            ss.stop_flag.clear()
            st.rerun()

    # Run History
    if ss.run_history:
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
        with st.expander(f"Run History ({len(ss.run_history)} runs)", expanded=False):
            for i, h in enumerate(reversed(ss.run_history)):
                status_icon = "\u26a0\ufe0f" if h.get("stopped") else "\u2705"
                st.markdown(
                    f'<div class="history-row">'
                    f'<span>{status_icon} #{len(ss.run_history) - i}</span>'
                    f'<span class="h-date">{h["date"]}</span>'
                    f'<span class="h-records">{h["records"]:,} / {h["total"]:,} records</span>'
                    f'<span class="h-priority">{h["priority"]:,} priority</span>'
                    f'<span class="h-time">{h["elapsed"]}</span>'
                    f'</div>',
                    unsafe_allow_html=True)
            if st.button("Clear History"):
                ss.run_history = []
                st.rerun()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB: DATABASE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif active_tab == "database":
    stats = db_get_total_stats()

    st.markdown(f'''
    <div class="dash-stats">
        <div class="dash-stat">
            <div class="ds-label">Total Runs</div>
            <div class="ds-value ds-blue">{stats["total_runs"]}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Total Records</div>
            <div class="ds-value ds-green">{stats["total_records"]:,}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Priority Matches</div>
            <div class="ds-value ds-purple">{stats["total_priority"]:,}</div>
        </div>
        <div class="dash-stat">
            <div class="ds-label">Total Errors</div>
            <div class="ds-value ds-red">{stats["total_errors"]:,}</div>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    if stats["total_runs"] == 0:
        st.info("No data in the database yet. Run a scraping job to start collecting data.")
    else:
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        view_mode = st.radio(
            "View Mode", ["All Data Combined", "By Run"],
            horizontal=True, key="db_view_mode")

        if view_mode == "All Data Combined":
            all_results = db_get_all_results()
            if all_results:
                df = pd.DataFrame(all_results)
                display_cols = ["Stock Number"] + [
                    c for c in df.columns
                    if c not in ("Stock Number", "_run_id", "_date", "_save_name")
                ] + ["_date", "_save_name"]
                display_cols = [c for c in display_cols if c in df.columns]
                df = df[display_cols]
                df = df.rename(columns={"_date": "Date", "_save_name": "File Name"})
                st.dataframe(df, use_container_width=True, height=400)

                dl_all1, dl_all2 = st.columns(2)
                with dl_all1:
                    csv_all = df.to_csv(index=False).encode("utf-8")
                    st.download_button(
                        "Download All Data (CSV)",
                        data=csv_all,
                        file_name=f"ALL_DATA_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True)
                with dl_all2:
                    xlsx_buf = io.BytesIO()
                    df.to_excel(xlsx_buf, index=False, engine="openpyxl")
                    xlsx_buf.seek(0)
                    st.download_button(
                        "Download All Data (Excel)",
                        data=xlsx_buf.getvalue(),
                        file_name=f"ALL_DATA_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
        else:
            all_runs = db_get_all_runs()
            for run_item in all_runs:
                status_icon = "\u26a0\ufe0f" if run_item["was_stopped"] else "\u2705"
                run_label = (
                    f'{status_icon} {run_item["save_name"]} | '
                    f'{run_item["processed"]:,} records | '
                    f'{run_item["created_at"][:16]}'
                )
                with st.expander(run_label, expanded=False):
                    rc1, rc2, rc3, rc4 = st.columns(4)
                    rc1.metric("Processed", f"{run_item['processed']:,}")
                    rc2.metric("Priority", f"{run_item['priority_count']:,}")
                    rc3.metric("Blacklisted", f"{run_item['blacklisted']:,}")
                    rc4.metric("Elapsed", run_item["elapsed"])

                    run_results = db_get_run_results(run_item["run_id"])
                    if run_results:
                        run_df = pd.DataFrame(run_results)
                        st.dataframe(run_df, use_container_width=True, height=300)

                        csv_run = run_df.to_csv(index=False).encode("utf-8")
                        st.download_button(
                            f"Download {run_item['save_name']} (CSV)",
                            data=csv_run,
                            file_name=f"{run_item['save_name']}_{run_item['run_id']}.csv",
                            mime="text/csv",
                            key=f"dl_{run_item['run_id']}",
                            use_container_width=True)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
        if st.button("Clear All Database", key="clear_db"):
            db_clear_all()
            st.rerun()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB: SETTINGS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
elif active_tab == "settings":
    st.markdown('<div class="section-label">Appearance</div>', unsafe_allow_html=True)
    current_theme = "Dark Mode" if is_dark else "Light Mode"
    st.markdown(f'''
    <div class="p-card">
        <div style="display:flex; justify-content:space-between; align-items:center;">
            <div>
                <div style="font-weight:700; font-size:0.95rem;">Theme: {current_theme}</div>
                <div style="font-size:0.78rem; color:{MUTED} !important; margin-top:4px;">
                    Toggle between dark and light themes using the button in the top right corner.
                </div>
            </div>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    st.markdown('<div class="section-label">System Information</div>', unsafe_allow_html=True)
    st.markdown(f'''
    <div class="sys-bar">
        <div class="sys-item">
            <span class="si-label">Platform</span>
            <span class="si-value">{os.name}</span>
        </div>
        <div class="sys-item">
            <span class="si-label">Python</span>
            <span class="si-value">{os.sys.version.split()[0] if hasattr(os, 'sys') else 'N/A'}</span>
        </div>
        <div class="sys-item">
            <span class="si-label">Selenium</span>
            <span class="si-value" style="color:{GREEN2 if SELENIUM_OK else RED2} !important">
                {"Installed" if SELENIUM_OK else "Not Found"}
            </span>
        </div>
        <div class="sys-item">
            <span class="si-label">psutil</span>
            <span class="si-value" style="color:{GREEN2 if PSUTIL_OK else RED2} !important">
                {"Installed" if PSUTIL_OK else "Not Found"}
            </span>
        </div>
    </div>
    ''', unsafe_allow_html=True)

    st.markdown('<div class="section-label">About</div>', unsafe_allow_html=True)
    st.markdown(f'''
    <div class="p-card" style="text-align:center;">
        <div style="font-size:1.5rem; font-weight:900; margin-bottom:0.5rem;">3LINES DataHunter</div>
        <div style="font-size:0.85rem; color:{MUTED} !important;">
            v14.0 Premium Edition<br>
            Real-Time Resource Auditing &bull; Dual Filtering &bull; Priority Targets &bull; Blacklist Exclusion<br>
            Multi-Format Export &bull; Auto-Retry &bull; Performance Analytics
        </div>
    </div>
    ''', unsafe_allow_html=True)


# ── Footer ──
st.markdown(f'''
<div class="app-footer">
    3LINES DataHunter v14.0 Premium &mdash; Real-Time Resource Auditing &bull;
    Dual Filtering &bull; Priority Targets &bull; Blacklist Exclusion
</div>
''', unsafe_allow_html=True)
