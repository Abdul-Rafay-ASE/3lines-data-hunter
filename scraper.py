import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
import os
from datetime import datetime

# ============================================
# CONFIGURATION - Modify these as needed
# ============================================
BLACKLIST = ["MILITARY", "FEDERAL", "FINLAND", "A486G"]  # Exclude these companies
PREFERRED_VENDOR = "AMETEK"  # Always first priority
INPUT_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "Abdullah", "input.xlsx")
OUTPUT_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "Abdullah")
# Use a session URL - user needs to provide a valid session URL or the script uses this default
BASE_URL = "https://www.lqlite.com/Lq_FLIS.aspx?B=JTEyVFAwJTIyWiUyNDFTJTBEfiU3QyUwQWMlMTJVWlglMEQlMDR3cyUwRHVzJTA3JTAwJTBDJTA4cCUwOSUwNyZTRT00MDFEQzhFQjYwODBDMDQ0Ng=="

# Color definitions - Professional colors
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # AMETEK (Priority 1) - Bright green
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Part Number Match (Priority 2) - Soft yellow
HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Header - Dark blue
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")   # Title - Navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14, name='Calibri')
DATA_FONT = Font(size=10, name='Calibri')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)
MEDIUM_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)


def read_input_file(file_path):
    """Read input Excel file with Stock_Number and Part_Number columns"""
    if not os.path.exists(file_path):
        print(f"ERROR: Input file not found: {file_path}")
        print(f"Please create an Excel file with columns: Stock_Number, Part_Number")
        return []

    df = pd.read_excel(file_path)

    # Check for required columns (flexible naming)
    df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

    if 'stock_number' not in df.columns or 'part_number' not in df.columns:
        # Try first two columns if headers don't match
        if len(df.columns) >= 2:
            df.columns = ['stock_number', 'part_number'] + list(df.columns[2:])
        else:
            print("ERROR: Excel must have at least 2 columns (Stock_Number, Part_Number)")
            return []

    items = df[['stock_number', 'part_number']].dropna().values.tolist()
    print(f"Loaded {len(items)} items from: {file_path}")
    return items


def search_and_scrape(driver, search_term, search_type="nsn"):
    """Search by NSN or Part Number and scrape results"""
    try:
        driver.get(BASE_URL)
        time.sleep(5)

        # Use JavaScript to set the input value and trigger search
        input_id = "nNIIN" if search_type == "nsn" else "nPART"

        # Try to set value using JavaScript
        js_script = f"""
            var input = document.getElementById('{input_id}');
            if (input) {{
                input.value = '{search_term}';
                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                input.dispatchEvent(new Event('change', {{ bubbles: true }}));
                return true;
            }}
            return false;
        """
        result = driver.execute_script(js_script)

        if not result:
            print(f"  Could not find input {input_id}")
            return []

        # Try to submit by pressing Enter or clicking a search button
        try:
            search_input = driver.find_element(By.ID, input_id)
            search_input.send_keys(Keys.RETURN)
        except:
            # Try clicking any search/submit button
            try:
                driver.execute_script(f"document.getElementById('{input_id}').form.submit();")
            except:
                pass

        time.sleep(6)

        # Scrape table data - look for the Part Information table specifically
        rows = driver.find_elements(By.TAG_NAME, "tr")
        raw_data = []
        in_part_table = False

        for row in rows:
            cells = row.find_elements(By.TAG_NAME, "td")
            row_text = row.text.upper()

            # Skip metadata rows (Item Name, Assignment Date, etc.)
            if "ITEM NAME:" in row_text or "ASSIGNMENT DATE:" in row_text:
                continue
            if "DATE STANDARDIZED:" in row_text or "CANCELLATION DATE:" in row_text:
                continue
            if "SCHEDULE B:" in row_text or "ORIGINATOR:" in row_text:
                continue
            if "NIIN:" in row_text or "FSC:" in row_text or "ISC:" in row_text:
                continue
            if "NSC:" in row_text or "ESD:" in row_text or "TIIC:" in row_text:
                continue

            # Check for header row to start capturing
            if "PART NUMBER" in row_text and "CAGE" in row_text and "COMPANY" in row_text:
                in_part_table = True
                continue  # Skip the header row itself

            if len(cells) > 5:
                row_content = [c.text for c in cells]
                first_cell = row_content[0] if row_content else ""

                # Skip if first cell looks like metadata
                if any(x in first_cell.upper() for x in ["NIIN:", "FSC:", "NSC:", "ITEM NAME", "DATE"]):
                    continue

                # Filter out blacklisted companies
                row_text_full = "".join(row_content).upper()
                if not any(word in row_text_full for word in BLACKLIST):
                    raw_data.append(row_content)

        return raw_data

    except Exception as e:
        print(f"  Search error: {e}")
        return []


def get_priority(supplier_part, company_name, user_part_number):
    """
    Determine priority for sorting:
    0 = AMETEK (Green) - Always first
    1 = Part Number Match (Yellow) - Company's part matches user's part
    2 = Standard (White) - All others
    """
    company_upper = str(company_name).upper()
    supplier_part_upper = str(supplier_part).upper()
    user_part_upper = str(user_part_number).upper()

    # Priority 1: AMETEK always first
    if PREFERRED_VENDOR.upper() in company_upper:
        return 0, "ametek"

    # Priority 2: Part number match
    if user_part_upper and user_part_upper in supplier_part_upper:
        return 1, "part_match"

    # Priority 3: Standard
    return 2, "standard"


def process_batch(driver, items):
    """Process all items from input file"""
    all_results = []
    total = len(items)

    for idx, (stock_number, part_number) in enumerate(items, 1):
        print(f"\n[{idx}/{total}] Processing NSN: {stock_number}, Part: {part_number}")

        # Try NSN search first
        results = search_and_scrape(driver, stock_number, "nsn")

        # If no results, try Part Number search
        if not results:
            print(f"  No results for NSN, trying Part Number...")
            results = search_and_scrape(driver, part_number, "part")

        if results:
            print(f"  Found {len(results)} suppliers")
            # Add metadata to each result
            for row in results:
                # Add: User's NSN, User's Part Number at the beginning
                enriched_row = [stock_number, part_number] + row
                all_results.append(enriched_row)
        else:
            print(f"  No results found")

        # Small delay between searches to avoid rate limiting
        time.sleep(1)

    return all_results


def create_formatted_excel(results, output_path):
    """Create Excel file with HORIZONTAL layout - one row per NSN, suppliers spread horizontally"""
    if not results:
        print("No results to export")
        return None

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Supplier Data"

    # Group results by NSN
    from collections import defaultdict
    nsn_groups = defaultdict(list)

    for row in results:
        user_nsn = row[0]
        user_part = row[1]
        supplier_part = row[2] if len(row) > 2 else ""
        company = row[4] if len(row) > 4 else ""

        priority_num, priority_type = get_priority(supplier_part, company, user_part)
        nsn_groups[(user_nsn, user_part)].append({
            'company': company,
            'part': supplier_part,
            'priority': priority_num,
            'type': priority_type
        })

    # Sort suppliers within each NSN group (AMETEK first, then part_match, then standard)
    for key in nsn_groups:
        nsn_groups[key].sort(key=lambda x: x['priority'])

    # Find max number of suppliers for any NSN (to determine column count)
    max_suppliers = max(len(suppliers) for suppliers in nsn_groups.values()) if nsn_groups else 1

    # === BUILD HEADERS ===
    # User_NSN | User_Part | OEM1 | PN1 | OEM2 | PN2 | OEM3 | PN3 | ...
    headers = ["User_NSN", "User_Part"]
    for i in range(1, max_suppliers + 1):
        headers.extend([f"OEM {i}", f"PN {i}"])

    # === WRITE HEADER ROW ===
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws_data.row_dimensions[1].height = 25

    # === WRITE DATA ROWS ===
    row_idx = 2
    for (user_nsn, user_part), suppliers in nsn_groups.items():
        # Column A: User_NSN
        cell_nsn = ws_data.cell(row=row_idx, column=1, value=user_nsn)
        cell_nsn.border = THIN_BORDER
        cell_nsn.font = DATA_FONT

        # Column B: User_Part
        cell_part = ws_data.cell(row=row_idx, column=2, value=user_part)
        cell_part.border = THIN_BORDER
        cell_part.font = DATA_FONT

        # Columns C onwards: OEM1, PN1, OEM2, PN2, ...
        col_idx = 3
        for supplier in suppliers:
            # OEM (Company) column
            cell_oem = ws_data.cell(row=row_idx, column=col_idx, value=supplier['company'])
            cell_oem.border = THIN_BORDER
            cell_oem.font = DATA_FONT

            # Apply color based on priority
            if supplier['type'] == "ametek":
                cell_oem.fill = GREEN_FILL
            elif supplier['type'] == "part_match":
                cell_oem.fill = YELLOW_FILL

            # PN (Part Number) column
            cell_pn = ws_data.cell(row=row_idx, column=col_idx + 1, value=supplier['part'])
            cell_pn.border = THIN_BORDER
            cell_pn.font = DATA_FONT

            col_idx += 2

        row_idx += 1

    # === FREEZE HEADER ROW ===
    ws_data.freeze_panes = 'A2'

    # === COLUMN WIDTHS ===
    ws_data.column_dimensions['A'].width = 18  # NSN
    ws_data.column_dimensions['B'].width = 15  # User_Part

    # Set widths for OEM/PN columns
    for i in range(3, len(headers) + 1):
        col_letter = ws_data.cell(row=1, column=i).column_letter
        if (i - 3) % 2 == 0:  # OEM columns (odd positions after B)
            ws_data.column_dimensions[col_letter].width = 30
        else:  # PN columns
            ws_data.column_dimensions[col_letter].width = 15

    # === PRINT SETTINGS ===
    ws_data.page_setup.orientation = 'landscape'
    ws_data.page_setup.fitToPage = True
    ws_data.page_setup.fitToWidth = 1

    # ========== SHEET 2: Summary ==========
    ws_summary = wb.create_sheet(title="Summary")

    ws_summary.cell(row=1, column=1, value="PROCUREMENT SUMMARY REPORT").font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:C1')

    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Counts
    ametek_count = sum(1 for p in processed_results if p[1] == "ametek")
    part_match_count = sum(1 for p in processed_results if p[1] == "part_match")
    standard_count = sum(1 for p in processed_results if p[1] == "standard")
    unique_nsns = len(set(r[2][0] for r in processed_results))

    ws_summary.cell(row=4, column=1, value="Total NSNs Processed:").font = Font(bold=True)
    ws_summary.cell(row=4, column=2, value=unique_nsns)

    ws_summary.cell(row=5, column=1, value="Total Suppliers Found:").font = Font(bold=True)
    ws_summary.cell(row=5, column=2, value=len(processed_results))

    ws_summary.cell(row=7, column=1, value="PRIORITY BREAKDOWN").font = Font(bold=True, size=12)

    ws_summary.cell(row=8, column=1, value="AMETEK (Priority 1):").font = Font(bold=True)
    ws_summary.cell(row=8, column=2, value=ametek_count)
    ws_summary.cell(row=8, column=1).fill = GREEN_FILL
    ws_summary.cell(row=8, column=2).fill = GREEN_FILL

    ws_summary.cell(row=9, column=1, value="Part Number Match (Priority 2):").font = Font(bold=True)
    ws_summary.cell(row=9, column=2, value=part_match_count)
    ws_summary.cell(row=9, column=1).fill = YELLOW_FILL
    ws_summary.cell(row=9, column=2).fill = YELLOW_FILL

    ws_summary.cell(row=10, column=1, value="Standard (Priority 3):").font = Font(bold=True)
    ws_summary.cell(row=10, column=2, value=standard_count)

    # Blacklist info
    ws_summary.cell(row=12, column=1, value="Excluded (Blacklist):").font = Font(bold=True)
    ws_summary.cell(row=12, column=2, value=", ".join(BLACKLIST))

    # Legend
    ws_summary.cell(row=14, column=1, value="LEGEND").font = Font(bold=True, size=12)
    ws_summary.cell(row=15, column=1, value="Green = AMETEK (Always First)")
    ws_summary.cell(row=15, column=1).fill = GREEN_FILL
    ws_summary.cell(row=16, column=1, value="Yellow = Part Number Match")
    ws_summary.cell(row=16, column=1).fill = YELLOW_FILL
    ws_summary.cell(row=17, column=1, value="White = Standard Supplier")

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20

    # Save
    wb.save(output_path)
    return output_path


def main():
    """Main function - Batch process NSNs from Excel input"""
    print("="*60)
    print("FLIS PROCUREMENT SCRAPER - Batch Processing")
    print("="*60)

    # Create output folder
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    # Read input file
    items = read_input_file(INPUT_FILE)
    if not items:
        print("\nNo items to process. Please check your input file.")
        print(f"Expected location: {INPUT_FILE}")
        print("Expected columns: Stock_Number, Part_Number")
        return

    print(f"\nProcessing {len(items)} items...")
    print(f"Blacklist: {BLACKLIST}")
    print(f"Priority 1: {PREFERRED_VENDOR}")
    print("="*60)

    # Initialize browser
    driver = webdriver.Chrome()

    try:
        # Process all items
        results = process_batch(driver, items)

        if results:
            # Generate output filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(OUTPUT_FOLDER, f"Procurement_Report_{timestamp}.xlsx")

            # Create formatted Excel
            create_formatted_excel(results, output_path)

            print("\n" + "="*60)
            print("SUCCESS! Report generated:")
            print(f"  File: {output_path}")
            print(f"  Total suppliers: {len(results)}")
            print("="*60)

            # Open the file
            os.startfile(output_path)
        else:
            print("\nNo results found for any items.")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()
